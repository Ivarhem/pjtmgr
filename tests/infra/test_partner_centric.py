"""거래처 중심 구조 전환 E2E 테스트.

partner_id scope 조회, 기간 필터, PortMap nullable, Export 검증.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from app.modules.common.models.audit_log import AuditLog
from app.modules.common.models.contract import Contract
from app.modules.common.models.contract_period import ContractPeriod
from app.modules.common.models.partner import Partner
from app.modules.infra.models.asset import Asset
from app.modules.infra.models.ip_subnet import IpSubnet
from app.modules.infra.models.port_map import PortMap
from app.modules.infra.models.policy_assignment import PolicyAssignment
from app.modules.infra.models.period_deliverable import PeriodDeliverable
from app.modules.infra.models.period_phase import PeriodPhase
from app.modules.infra.models.period_asset import PeriodAsset
from app.modules.infra.services.asset_service import list_assets, create_asset
from app.modules.infra.services.infra_metrics import (
    get_non_compliant_assignments,
    get_unsubmitted_deliverables,
    list_audit_logs,
    list_periods_summary,
)
from app.modules.infra.services.network_service import list_subnets, list_port_maps
from app.modules.infra.services.infra_exporter import export_partner
from app.modules.infra.schemas.asset import AssetCreate


# -- Helpers --


def _make_admin(db_session, admin_role_id: int):
    from app.modules.common.models.user import User

    user = User(login_id="centric_admin", name="CentricAdmin", role_id=admin_role_id)
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


def _make_partner(db_session, name: str = "고객사A") -> Partner:
    c = Partner(name=name)
    db_session.add(c)
    db_session.commit()
    db_session.refresh(c)
    return c


def _make_period(db_session, partner_id: int, code: str = "PRD-001") -> ContractPeriod:
    contract = Contract(
        contract_name=f"사업 {code}",
        contract_type="인프라",
        end_partner_id=partner_id,
    )
    db_session.add(contract)
    db_session.flush()
    period = ContractPeriod(
        contract_id=contract.id,
        period_year=2025,
        period_label=f"Y25-{code}",
        stage="50%",
        partner_id=partner_id,
    )
    db_session.add(period)
    db_session.commit()
    db_session.refresh(period)
    return period


def _make_asset(db_session, partner_id: int, name: str, asset_type: str = "server") -> Asset:
    a = Asset(partner_id=partner_id, asset_name=name, asset_type=asset_type)
    db_session.add(a)
    db_session.commit()
    db_session.refresh(a)
    return a


def _link_period_asset(db_session, contract_period_id: int, asset_id: int) -> None:
    pa = PeriodAsset(contract_period_id=contract_period_id, asset_id=asset_id)
    db_session.add(pa)
    db_session.commit()


# -- partner scope 조회 테스트 --


def test_list_assets_by_partner(db_session, admin_role_id) -> None:
    """거래처 범위 자산 조회: 다른 거래처 자산은 조회되지 않음."""
    cust_a = _make_partner(db_session, "고객A")
    cust_b = _make_partner(db_session, "고객B")

    _make_asset(db_session, cust_a.id, "SVR-A1")
    _make_asset(db_session, cust_a.id, "SVR-A2")
    _make_asset(db_session, cust_b.id, "SVR-B1")

    assets_a = list_assets(db_session, partner_id=cust_a.id)
    assets_b = list_assets(db_session, partner_id=cust_b.id)

    assert len(assets_a) == 2
    assert len(assets_b) == 1
    assert all(a.partner_id == cust_a.id for a in assets_a)


def test_list_assets_with_period_filter(db_session, admin_role_id) -> None:
    """기간 필터 적용: PeriodAsset으로 연결된 자산만 반환."""
    cust = _make_partner(db_session)
    period = _make_period(db_session, cust.id, "PRD-FILTER")

    a1 = _make_asset(db_session, cust.id, "LINKED-SVR")
    a2 = _make_asset(db_session, cust.id, "UNLINKED-SVR")
    _link_period_asset(db_session, period.id, a1.id)

    # 기간 필터 없이: 전체 2개
    all_assets = list_assets(db_session, partner_id=cust.id)
    assert len(all_assets) == 2

    # 기간 필터 적용: 연결된 1개만
    filtered = list_assets(db_session, partner_id=cust.id, period_id=period.id)
    assert len(filtered) == 1
    assert filtered[0].asset_name == "LINKED-SVR"


# -- subnet scope 테스트 --


def test_list_subnets_by_partner(db_session, admin_role_id) -> None:
    """거래처 범위 서브넷 조회."""
    cust_a = _make_partner(db_session, "SubnetCustA")
    cust_b = _make_partner(db_session, "SubnetCustB")

    db_session.add(IpSubnet(partner_id=cust_a.id, name="서버대역", subnet="10.0.1.0/24"))
    db_session.add(IpSubnet(partner_id=cust_b.id, name="관리대역", subnet="10.0.2.0/24"))
    db_session.commit()

    subs_a = list_subnets(db_session, partner_id=cust_a.id)
    subs_b = list_subnets(db_session, partner_id=cust_b.id)

    assert len(subs_a) == 1
    assert subs_a[0].name == "서버대역"
    assert len(subs_b) == 1


# -- PortMap nullable 테스트 --


def test_portmap_nullable_assets(db_session, admin_role_id) -> None:
    """src_asset_id, dst_asset_id가 NULL인 포트맵 (외부 구간)."""
    cust = _make_partner(db_session, "PortMapCust")

    pm = PortMap(
        partner_id=cust.id,
        src_asset_id=None,
        dst_asset_id=None,
        src_hostname="EXT-GW",
        dst_hostname="INT-FW",
        protocol="TCP",
        port=443,
        status="required",
    )
    db_session.add(pm)
    db_session.commit()

    portmaps = list_port_maps(db_session, partner_id=cust.id)
    assert len(portmaps) >= 1
    ext = [p for p in portmaps if p.src_hostname == "EXT-GW"]
    assert len(ext) == 1
    assert ext[0].src_asset_id is None
    assert ext[0].dst_asset_id is None


# -- Export 테스트 --


def test_export_partner_generates_xlsx(db_session, admin_role_id) -> None:
    """거래처 단위 Export: 3개 시트 생성 확인."""
    cust = _make_partner(db_session, "ExportCust")
    _make_asset(db_session, cust.id, "EXP-SVR-01", "server")
    db_session.add(IpSubnet(partner_id=cust.id, name="EXP대역", subnet="10.1.0.0/24"))
    db_session.add(PortMap(partner_id=cust.id, src_hostname="A", dst_hostname="B", status="required"))
    db_session.commit()

    xlsx_bytes = export_partner(db_session, cust.id)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))

    assert "01. Inventory" in wb.sheetnames
    assert "05. 네트워크 대역" in wb.sheetnames
    assert "03. Portmap" in wb.sheetnames

    # Inventory 시트에 데이터 존재 확인
    ws = wb["01. Inventory"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) >= 1


def test_export_partner_with_period_filter(db_session, admin_role_id) -> None:
    """기간 필터 적용 Export: 연결된 자산만 포함."""
    cust = _make_partner(db_session, "FilterExportCust")
    period = _make_period(db_session, cust.id, "EXP-PRD")

    a1 = _make_asset(db_session, cust.id, "PERIOD-ASSET")
    a2 = _make_asset(db_session, cust.id, "OTHER-ASSET")
    _link_period_asset(db_session, period.id, a1.id)

    xlsx_bytes = export_partner(db_session, cust.id, period_id=period.id)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb["01. Inventory"]

    # 헤더 제외, 데이터 행만 확인
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v for v in r)]
    assert len(data_rows) == 1  # PERIOD-ASSET만 포함


def test_dashboard_metrics_respect_partner_scope(db_session, admin_role_id) -> None:
    """현황판 집계 서비스는 partner_id 필터를 적용한다."""
    cust_a = _make_partner(db_session, "대시보드고객A")
    cust_b = _make_partner(db_session, "대시보드고객B")
    period_a = _make_period(db_session, cust_a.id, "DASH-A")
    period_b = _make_period(db_session, cust_b.id, "DASH-B")

    phase_a = PeriodPhase(contract_period_id=period_a.id, phase_type="build", status="in_progress")
    phase_b = PeriodPhase(contract_period_id=period_b.id, phase_type="test", status="in_progress")
    db_session.add_all([phase_a, phase_b])
    db_session.flush()

    db_session.add_all(
        [
            PeriodDeliverable(period_phase_id=phase_a.id, name="A-산출물", is_submitted=False),
            PeriodDeliverable(period_phase_id=phase_b.id, name="B-산출물", is_submitted=False),
        ]
    )
    db_session.commit()

    summary = list_periods_summary(db_session, partner_id=cust_a.id)
    unsubmitted = get_unsubmitted_deliverables(db_session, partner_id=cust_a.id)

    assert len(summary) == 1
    assert len(unsubmitted) == 1


def test_non_compliant_assignments_respect_partner_scope(db_session, admin_role_id) -> None:
    """미준수 정책 목록은 partner_id 필터를 적용한다."""
    cust_a = _make_partner(db_session, "정책고객A")
    cust_b = _make_partner(db_session, "정책고객B")
    asset_a = _make_asset(db_session, cust_a.id, "POL-A")
    asset_b = _make_asset(db_session, cust_b.id, "POL-B")

    db_session.add_all(
        [
            PolicyAssignment(partner_id=cust_a.id, asset_id=asset_a.id, status="non_compliant"),
            PolicyAssignment(partner_id=cust_b.id, asset_id=asset_b.id, status="non_compliant"),
        ]
    )
    db_session.commit()

    rows = get_non_compliant_assignments(db_session, partner_id=cust_a.id)

    assert len(rows) == 1
    assert rows[0]["partner_id"] == cust_a.id


def test_list_audit_logs_returns_enriched_user_name(db_session, admin_role_id) -> None:
    """감사로그 조회는 service에서 사용자명 enrichment를 수행한다."""
    admin = _make_admin(db_session, admin_role_id)
    db_session.add(
        AuditLog(
            user_id=admin.id,
            action="create",
            entity_type="contract_period",
            entity_id=1,
            module="infra",
            summary="기간 생성",
        )
    )
    db_session.commit()

    rows = list_audit_logs(db_session, module="infra")

    assert len(rows) == 1
    assert rows[0]["user_name"] == "CentricAdmin"
