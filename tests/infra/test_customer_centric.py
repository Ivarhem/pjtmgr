"""고객사 중심 구조 전환 E2E 테스트.

customer_id scope 조회, 프로젝트 필터, PortMap nullable, Export 검증.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from app.modules.common.models.customer import Customer
from app.modules.infra.models.asset import Asset
from app.modules.infra.models.ip_subnet import IpSubnet
from app.modules.infra.models.port_map import PortMap
from app.modules.infra.models.project import Project
from app.modules.infra.models.project_asset import ProjectAsset
from app.modules.infra.services.asset_service import list_assets, create_asset
from app.modules.infra.services.network_service import list_subnets, list_port_maps
from app.modules.infra.services.infra_exporter import export_customer
from app.modules.infra.schemas.asset import AssetCreate


# ── Helpers ──


def _make_admin(db_session, admin_role_id: int):
    from app.modules.common.models.user import User

    user = User(login_id="centric_admin", name="CentricAdmin", role_id=admin_role_id)
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


def _make_customer(db_session, name: str = "고객사A") -> Customer:
    c = Customer(name=name)
    db_session.add(c)
    db_session.commit()
    db_session.refresh(c)
    return c


def _make_project(db_session, customer_id: int, code: str = "PRJ-001") -> Project:
    p = Project(project_code=code, project_name=f"프로젝트 {code}", customer_id=customer_id)
    db_session.add(p)
    db_session.commit()
    db_session.refresh(p)
    return p


def _make_asset(db_session, customer_id: int, name: str, asset_type: str = "server") -> Asset:
    a = Asset(customer_id=customer_id, asset_name=name, asset_type=asset_type)
    db_session.add(a)
    db_session.commit()
    db_session.refresh(a)
    return a


def _link_project_asset(db_session, project_id: int, asset_id: int) -> None:
    pa = ProjectAsset(project_id=project_id, asset_id=asset_id)
    db_session.add(pa)
    db_session.commit()


# ── customer scope 조회 테스트 ──


def test_list_assets_by_customer(db_session, admin_role_id) -> None:
    """고객사 범위 자산 조회: 다른 고객사 자산은 조회되지 않음."""
    cust_a = _make_customer(db_session, "고객A")
    cust_b = _make_customer(db_session, "고객B")

    _make_asset(db_session, cust_a.id, "SVR-A1")
    _make_asset(db_session, cust_a.id, "SVR-A2")
    _make_asset(db_session, cust_b.id, "SVR-B1")

    assets_a = list_assets(db_session, customer_id=cust_a.id)
    assets_b = list_assets(db_session, customer_id=cust_b.id)

    assert len(assets_a) == 2
    assert len(assets_b) == 1
    assert all(a.customer_id == cust_a.id for a in assets_a)


def test_list_assets_with_project_filter(db_session, admin_role_id) -> None:
    """프로젝트 필터 적용: ProjectAsset으로 연결된 자산만 반환."""
    cust = _make_customer(db_session)
    proj = _make_project(db_session, cust.id, "PRJ-FILTER")

    a1 = _make_asset(db_session, cust.id, "LINKED-SVR")
    a2 = _make_asset(db_session, cust.id, "UNLINKED-SVR")
    _link_project_asset(db_session, proj.id, a1.id)

    # 프로젝트 필터 없이: 전체 2개
    all_assets = list_assets(db_session, customer_id=cust.id)
    assert len(all_assets) == 2

    # 프로젝트 필터 적용: 연결된 1개만
    filtered = list_assets(db_session, customer_id=cust.id, project_id=proj.id)
    assert len(filtered) == 1
    assert filtered[0].asset_name == "LINKED-SVR"


# ── subnet scope 테스트 ──


def test_list_subnets_by_customer(db_session, admin_role_id) -> None:
    """고객사 범위 서브넷 조회."""
    cust_a = _make_customer(db_session, "SubnetCustA")
    cust_b = _make_customer(db_session, "SubnetCustB")

    db_session.add(IpSubnet(customer_id=cust_a.id, name="서버대역", subnet="10.0.1.0/24"))
    db_session.add(IpSubnet(customer_id=cust_b.id, name="관리대역", subnet="10.0.2.0/24"))
    db_session.commit()

    subs_a = list_subnets(db_session, customer_id=cust_a.id)
    subs_b = list_subnets(db_session, customer_id=cust_b.id)

    assert len(subs_a) == 1
    assert subs_a[0].name == "서버대역"
    assert len(subs_b) == 1


# ── PortMap nullable 테스트 ──


def test_portmap_nullable_assets(db_session, admin_role_id) -> None:
    """src_asset_id, dst_asset_id가 NULL인 포트맵 (외부 구간)."""
    cust = _make_customer(db_session, "PortMapCust")

    pm = PortMap(
        customer_id=cust.id,
        src_asset_id=None,
        dst_asset_id=None,
        src_hostname="EXT-GW",
        dst_hostname="INT-FW",
        protocol="TCP",
        port=443,
        status="required",
    )
    db_session.add(pm)
    db_session.commit()

    portmaps = list_port_maps(db_session, customer_id=cust.id)
    assert len(portmaps) >= 1
    ext = [p for p in portmaps if p.src_hostname == "EXT-GW"]
    assert len(ext) == 1
    assert ext[0].src_asset_id is None
    assert ext[0].dst_asset_id is None


# ── Export 테스트 ──


def test_export_customer_generates_xlsx(db_session, admin_role_id) -> None:
    """고객사 단위 Export: 3개 시트 생성 확인."""
    cust = _make_customer(db_session, "ExportCust")
    _make_asset(db_session, cust.id, "EXP-SVR-01", "server")
    db_session.add(IpSubnet(customer_id=cust.id, name="EXP대역", subnet="10.1.0.0/24"))
    db_session.add(PortMap(customer_id=cust.id, src_hostname="A", dst_hostname="B", status="required"))
    db_session.commit()

    xlsx_bytes = export_customer(db_session, cust.id)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))

    assert "01. Inventory" in wb.sheetnames
    assert "05. 네트워크 대역" in wb.sheetnames
    assert "03. Portmap" in wb.sheetnames

    # Inventory 시트에 데이터 존재 확인
    ws = wb["01. Inventory"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) >= 1


def test_export_customer_with_project_filter(db_session, admin_role_id) -> None:
    """프로젝트 필터 적용 Export: 연결된 자산만 포함."""
    cust = _make_customer(db_session, "FilterExportCust")
    proj = _make_project(db_session, cust.id, "EXP-PROJ")

    a1 = _make_asset(db_session, cust.id, "PROJ-ASSET")
    a2 = _make_asset(db_session, cust.id, "OTHER-ASSET")
    _link_project_asset(db_session, proj.id, a1.id)

    xlsx_bytes = export_customer(db_session, cust.id, project_id=proj.id)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb["01. Inventory"]

    # 헤더 제외, 데이터 행만 확인
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v for v in r)]
    assert len(data_rows) == 1  # PROJ-ASSET만 포함
