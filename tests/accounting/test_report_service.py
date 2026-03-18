from io import BytesIO

from openpyxl import load_workbook

from app.modules.accounting.models.contract import Contract
from app.modules.accounting.models.contract_period import ContractPeriod
from app.modules.accounting.models.contract_type_config import ContractTypeConfig
from app.modules.common.models.customer import Customer
from app.modules.accounting.models.monthly_forecast import MonthlyForecast
from app.modules.accounting.models.receipt import Receipt
from app.modules.accounting.models.receipt_match import ReceiptMatch
from app.modules.accounting.models.transaction_line import STATUS_CONFIRMED, TransactionLine
from app.modules.common.models.user import User
from app.modules.accounting.services import report as report_service
from app.modules.accounting.services.metrics import build_filter


def _seed_report_data(db_session, user_role_id) -> dict[str, int]:
    owner = User(name="영업", login_id="sales-report", role_id=user_role_id, department="영업1팀")
    end_customer = Customer(name="엔드고객")
    billing_customer = Customer(name="매출처")
    db_session.add_all(
        [
            owner,
            end_customer,
            billing_customer,
            ContractTypeConfig(code="MA", label="MA", sort_order=1, is_active=True),
        ]
    )
    db_session.flush()

    contract = Contract(
        contract_name="보고서 테스트 사업",
        contract_type="MA",
        contract_code="MA-2026-0001",
        owner_user_id=owner.id,
        end_customer_id=end_customer.id,
        status="active",
    )
    db_session.add(contract)
    db_session.flush()

    period = ContractPeriod(
        contract_id=contract.id,
        period_year=2026,
        period_label="Y26",
        stage="70%",
        is_planned=True,
        start_month="2026-01-01",
        end_month="2026-12-01",
        owner_user_id=owner.id,
        customer_id=billing_customer.id,
    )
    db_session.add(period)
    db_session.flush()

    db_session.add_all(
        [
            MonthlyForecast(
                contract_period_id=period.id,
                forecast_month="2026-01-01",
                revenue_amount=120,
                gp_amount=90,
                version_no=1,
                is_current=True,
                created_by=owner.id,
            ),
            TransactionLine(
                contract_id=contract.id,
                revenue_month="2026-01-01",
                line_type="revenue",
                customer_id=billing_customer.id,
                supply_amount=100,
                status=STATUS_CONFIRMED,
                created_by=owner.id,
            ),
            TransactionLine(
                contract_id=contract.id,
                revenue_month="2026-01-01",
                line_type="cost",
                customer_id=billing_customer.id,
                supply_amount=30,
                status=STATUS_CONFIRMED,
                created_by=owner.id,
            ),
        ]
    )
    db_session.flush()

    january_receipt = Receipt(
        contract_id=contract.id,
        customer_id=billing_customer.id,
        receipt_date="2026-01-15",
        revenue_month="2026-01-01",
        amount=100,
        created_by=owner.id,
    )
    february_receipt = Receipt(
        contract_id=contract.id,
        customer_id=billing_customer.id,
        receipt_date="2026-02-10",
        revenue_month="2026-02-01",
        amount=30,
        created_by=owner.id,
    )
    db_session.add_all([january_receipt, february_receipt])
    db_session.flush()

    revenue_line = (
        db_session.query(TransactionLine)
        .filter(TransactionLine.contract_id == contract.id, TransactionLine.line_type == "revenue")
        .one()
    )
    db_session.add(
        ReceiptMatch(
            receipt_id=january_receipt.id,
            transaction_line_id=revenue_line.id,
            matched_amount=60,
            match_type="manual",
            created_by=owner.id,
        )
    )
    db_session.commit()

    return {"contract_id": contract.id}


def test_summary_forecast_vs_actual_and_receivables_use_expected_totals(db_session, user_role_id) -> None:
    ids = _seed_report_data(db_session, user_role_id)
    filt = build_filter("2026-01", "2026-02")

    summary = report_service.get_summary(db_session, filt)
    forecast_vs_actual = report_service.list_forecast_vs_actual(db_session, filt)
    receivables = report_service.list_receivables(db_session, filt)

    assert summary["kpis"] == {
        "forecast_revenue": 120,
        "actual_revenue": 100,
        "cost": 30,
        "gp": 70,
        "gp_pct": 70.0,
        "receipt": 130,
        "ar": 40,
        "achievement_rate": 83.3,
    }
    assert summary["period_summary"] == [
        {
            "month": "2026-01",
            "forecast_revenue": 120,
            "actual_revenue": 100,
            "cost": 30,
            "gp": 70,
            "gp_pct": 70.0,
            "receipt": 100,
            "ar": 40,
        },
        {
            "month": "2026-02",
            "forecast_revenue": 0,
            "actual_revenue": 0,
            "cost": 0,
            "gp": 0,
            "gp_pct": None,
            "receipt": 30,
            "ar": 0,
        },
    ]

    assert forecast_vs_actual["rows"] == [
        {
            "contract_id": ids["contract_id"],
            "contract_period_id": 1,
            "contract_name": "보고서 테스트 사업",
            "contract_type": "MA",
            "owner_name": "영업",
            "department": "영업1팀",
            "end_customer_name": "엔드고객",
            "stage": "70%",
            "forecast_revenue": 120,
            "actual_revenue": 100,
            "gap_revenue": 20,
            "achievement_rate": 83.3,
            "gp": 70,
            "gp_pct": 70.0,
        }
    ]
    assert forecast_vs_actual["totals"]["gap_revenue"] == 20
    assert receivables["rows"][0]["receipt"] == 130
    assert receivables["rows"][0]["ar"] == 40
    assert receivables["totals"] == {
        "actual_revenue": 100,
        "receipt": 130,
        "ar": 40,
        "ar_rate": 40.0,
    }


def test_contract_pnl_uses_matched_amount_for_ar(db_session, user_role_id) -> None:
    ids = _seed_report_data(db_session, user_role_id)

    result = report_service.get_contract_pnl(db_session, ids["contract_id"], 2026)

    assert result["months"] == ["2026-01-01", "2026-02-01"]
    assert result["receipt_totals"] == {
        "2026-01-01": 100,
        "2026-02-01": 30,
    }
    assert result["ar_monthly"] == {
        "2026-01-01": 40,
        "2026-02-01": 40,
    }
    assert result["grand_receipt"] == 130
    assert result["grand_ar"] == 40


def test_export_summary_writes_totals_row(db_session, user_role_id) -> None:
    _seed_report_data(db_session, user_role_id)
    filt = build_filter("2026-01", "2026-02")

    content = report_service.export_summary(db_session, filt)
    ws = load_workbook(BytesIO(content)).active

    assert ws.title == "요약 현황"
    assert ws["A5"].value == "2026-01"
    assert ws["H5"].value == 40
    assert ws["A7"].value == "합계"
    assert ws["B7"].value == 120
    assert ws["H7"].value == 40


def test_export_contract_pnl_writes_ar_from_matched_amount(db_session, user_role_id) -> None:
    ids = _seed_report_data(db_session, user_role_id)

    content = report_service.export_contract_pnl(db_session, ids["contract_id"], 2026)
    ws = load_workbook(BytesIO(content)).active

    target_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == "미수금":
            target_row = row
            break

    assert target_row is not None
    assert ws.cell(row=target_row, column=6).value == 40
    assert ws.cell(row=target_row, column=7).value == 40
    assert ws.cell(row=target_row, column=8).value == 40
