"""보고서 집계 서비스.

보고서 1: 요약 현황 (KPI + 월별 추이)
보고서 2: Forecast vs Actual (사업별 비교)
보고서 3: 미수 현황 (사업별 AR)
보고서 4: 매입매출관리 (사업별 거래처 피벗) — 기존 유지
"""
from __future__ import annotations

import datetime
from typing import TYPE_CHECKING

from sqlalchemy import func, and_
from sqlalchemy.orm import Session, joinedload

from app.models.contract import Contract
from app.models.contract_period import ContractPeriod
from app.models.monthly_forecast import MonthlyForecast
from app.models.transaction_line import TransactionLine, STATUS_CONFIRMED
from app.models.receipt import Receipt
from app.models.receipt_match import ReceiptMatch
from app.exceptions import NotFoundError
from app.schemas.report import ReportFilter
from app.services.metrics import (
    build_filter as _build_filter,
    month_range as _month_range,
    years_in_range as _years_in_range,
    apply_common_filters as _apply_common_filters,
    load_forecasts as _load_forecasts,
    load_actuals as _load_actuals,
    load_matched_totals as _load_matched_totals,
    load_matched_totals_monthly as _load_matched_totals_monthly,
    load_receipts as _load_receipts,
    load_receipts_monthly as _load_receipts_monthly,
    safe_pct as _safe_pct,
)

if TYPE_CHECKING:
    from app.models.user import User


# ═══ 보고서 1: 요약 현황 ═════════════════════════════════════════


def get_summary(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> dict:
    """요약 현황: KPI + 월별 추이 테이블."""
    q = (
        db.query(ContractPeriod)
        .join(ContractPeriod.contract)
        .options(joinedload(ContractPeriod.contract), joinedload(ContractPeriod.owner))
        .filter(Contract.status != "cancelled")
    )
    q, years = _apply_common_filters(q, filt, current_user)
    periods = q.all()

    months = _month_range(filt.date_from, filt.date_to)

    if not periods:
        return {
            "kpis": {"forecast_revenue": 0, "actual_revenue": 0, "cost": 0, "gp": 0,
                     "gp_pct": None, "receipt": 0, "ar": 0, "achievement_rate": None},
            "period_summary": [
                {"month": m[:7], "forecast_revenue": 0, "actual_revenue": 0, "cost": 0,
                 "gp": 0, "gp_pct": None, "receipt": 0, "ar": 0}
                for m in months
            ],
        }

    contract_ids = list({p.contract_id for p in periods})
    period_ids = [p.id for p in periods]

    fc_map = _load_forecasts(db, period_ids, filt.date_from, filt.date_to)
    act_map = _load_actuals(db, contract_ids, filt.date_from, filt.date_to)
    rcpt_monthly_map = _load_receipts_monthly(db, contract_ids, filt.date_from, filt.date_to)
    match_monthly_map = _load_matched_totals_monthly(db, contract_ids, filt.date_from, filt.date_to)

    # period_id → contract_id 매핑
    contract_period_ids: dict[int, list[int]] = {}
    for p in periods:
        contract_period_ids.setdefault(p.contract_id, []).append(p.id)

    monthly_rows: list[dict] = []
    total_fc_revenue = 0
    total_act_revenue = 0
    total_cost = 0
    total_payment = 0

    for m in months:
        fc_revenue = 0
        for did, pids in contract_period_ids.items():
            for pid in pids:
                fc = fc_map.get(pid, {}).get(m)
                if fc:
                    fc_revenue += fc["revenue"]

        act_revenue = 0
        act_cost = 0
        for did in contract_ids:
            ac = act_map.get(did, {}).get(m)
            if ac:
                act_revenue += ac["revenue"]
                act_cost += ac["cost"]

        pay = sum(rcpt_monthly_map.get(did, {}).get(m, 0) for did in contract_ids)
        allocated = sum(match_monthly_map.get(did, {}).get(m, 0) for did in contract_ids)
        ar = act_revenue - allocated

        gp = act_revenue - act_cost
        total_fc_revenue += fc_revenue
        total_act_revenue += act_revenue
        total_cost += act_cost
        total_payment += pay

        monthly_rows.append({
            "month": m[:7],
            "forecast_revenue": fc_revenue,
            "actual_revenue": act_revenue,
            "cost": act_cost,
            "gp": gp,
            "gp_pct": _safe_pct(gp, act_revenue),
            "receipt": pay,
            "ar": ar,
        })

    total_gp = total_act_revenue - total_cost
    kpis = {
        "forecast_revenue": total_fc_revenue,
        "actual_revenue": total_act_revenue,
        "cost": total_cost,
        "gp": total_gp,
        "gp_pct": _safe_pct(total_gp, total_act_revenue),
        "receipt": total_payment,
        "ar": sum(row["ar"] for row in monthly_rows),
        "achievement_rate": _safe_pct(total_act_revenue, total_fc_revenue),
    }

    return {"kpis": kpis, "period_summary": monthly_rows}


# ═══ 보고서 2: Forecast vs Actual ════════════════════════════════


def list_forecast_vs_actual(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> dict:
    """사업별 Forecast vs Actual 비교."""
    q = (
        db.query(ContractPeriod)
        .join(ContractPeriod.contract)
        .options(
            joinedload(ContractPeriod.contract).joinedload(Contract.end_customer),
            joinedload(ContractPeriod.contract).joinedload(Contract.owner),
            joinedload(ContractPeriod.owner),
        )
        .filter(Contract.status != "cancelled")
    )
    q, years = _apply_common_filters(q, filt, current_user)
    periods = q.order_by(Contract.contract_code).all()

    if not periods:
        return {"rows": [], "totals": None}

    contract_ids = list({p.contract_id for p in periods})
    period_ids = [p.id for p in periods]

    fc_map = _load_forecasts(db, period_ids, filt.date_from, filt.date_to)
    act_map = _load_actuals(db, contract_ids, filt.date_from, filt.date_to)

    # contract별 forecast 합산 (여러 period 가능)
    contract_fc: dict[int, int] = {}
    for p in periods:
        fc_months = fc_map.get(p.id, {})
        contract_fc[p.contract_id] = contract_fc.get(p.contract_id, 0) + sum(v["revenue"] for v in fc_months.values())

    rows: list[dict] = []
    seen_contracts: set[int] = set()
    total_fc = 0
    total_act_revenue = 0
    total_cost = 0

    for period in periods:
        contract = period.contract
        if contract.id in seen_contracts:
            continue
        seen_contracts.add(contract.id)
        owner = period.owner or contract.owner

        fc_revenue = contract_fc.get(contract.id, 0)
        ac_months = act_map.get(contract.id, {})
        act_revenue = sum(v["revenue"] for v in ac_months.values())
        act_cost = sum(v["cost"] for v in ac_months.values())

        gp = act_revenue - act_cost
        gap = fc_revenue - act_revenue

        total_fc += fc_revenue
        total_act_revenue += act_revenue
        total_cost += act_cost

        rows.append({
            "contract_id": contract.id,
            "contract_period_id": period.id,
            "contract_name": contract.contract_name,
            "contract_type": contract.contract_type,
            "owner_name": owner.name if owner else None,
            "department": owner.department if owner else None,
            "end_customer_name": contract.end_customer.name if contract.end_customer else None,
            "stage": period.stage,
            "forecast_revenue": fc_revenue,
            "actual_revenue": act_revenue,
            "gap_revenue": gap,
            "achievement_rate": _safe_pct(act_revenue, fc_revenue),
            "gp": gp,
            "gp_pct": _safe_pct(gp, act_revenue),
        })

    # gap 내림차순
    rows.sort(key=lambda r: r["gap_revenue"], reverse=True)

    total_gp = total_act_revenue - total_cost
    totals = {
        "contract_id": 0,
        "contract_period_id": 0,
        "contract_name": "합계",
        "contract_type": "",
        "owner_name": None,
        "department": None,
        "end_customer_name": None,
        "stage": None,
        "forecast_revenue": total_fc,
        "actual_revenue": total_act_revenue,
        "gap_revenue": total_fc - total_act_revenue,
        "achievement_rate": _safe_pct(total_act_revenue, total_fc),
        "gp": total_gp,
        "gp_pct": _safe_pct(total_gp, total_act_revenue),
    }

    return {"rows": rows, "totals": totals}


# ═══ 보고서 3: 미수 현황 ═════════════════════════════════════════


def list_receivables(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> dict:
    """사업별 미수금 현황."""
    q = (
        db.query(ContractPeriod)
        .join(ContractPeriod.contract)
        .options(
            joinedload(ContractPeriod.contract).joinedload(Contract.end_customer),
            joinedload(ContractPeriod.contract).joinedload(Contract.owner),
            joinedload(ContractPeriod.owner),
        )
        .filter(Contract.status != "cancelled")
    )
    q, years = _apply_common_filters(q, filt, current_user)
    periods = q.order_by(Contract.contract_code).all()

    if not periods:
        return {
            "rows": [],
            "totals": {"actual_revenue": 0, "receipt": 0, "ar": 0, "ar_rate": None},
        }

    contract_ids = list({p.contract_id for p in periods})
    act_map = _load_actuals(db, contract_ids, filt.date_from, filt.date_to)
    rcpt_map = _load_receipts(db, contract_ids, filt.date_from, filt.date_to)
    match_map = _load_matched_totals(db, contract_ids, filt.date_from, filt.date_to)

    rows: list[dict] = []
    seen_contracts: set[int] = set()
    total_act_revenue = 0
    total_receipt = 0
    total_allocated = 0

    for period in periods:
        contract = period.contract
        if contract.id in seen_contracts:
            continue
        seen_contracts.add(contract.id)
        owner = period.owner or contract.owner

        ac_months = act_map.get(contract.id, {})
        act_revenue = sum(v["revenue"] for v in ac_months.values())
        receipt = rcpt_map.get(contract.id, 0)
        allocated = match_map.get(contract.id, 0)
        ar = act_revenue - allocated

        if ar <= 0 and act_revenue == 0:
            continue

        total_act_revenue += act_revenue
        total_receipt += receipt
        total_allocated += allocated

        rows.append({
            "contract_id": contract.id,
            "contract_name": contract.contract_name,
            "contract_type": contract.contract_type,
            "owner_name": owner.name if owner else None,
            "department": owner.department if owner else None,
            "end_customer_name": contract.end_customer.name if contract.end_customer else None,
            "actual_revenue": act_revenue,
            "receipt": receipt,
            "ar": ar,
            "ar_rate": _safe_pct(ar, act_revenue),
        })

    rows.sort(key=lambda r: r["ar"], reverse=True)

    total_ar = total_act_revenue - total_allocated
    return {
        "rows": rows,
        "totals": {
            "actual_revenue": total_act_revenue,
            "receipt": total_receipt,
            "ar": total_ar,
            "ar_rate": _safe_pct(total_ar, total_act_revenue),
        },
    }


# ═══ 보고서 4: 매입매출관리 (기존 유지) ══════════════════════════


def get_contract_pnl(
    db: Session,
    contract_id: int,
    period_year: int | None = None,
) -> dict:
    """사업별 매입매출관리: 거래처별 월별 매출/매입/입금 피벗 + GP + 미수금."""
    contract = db.query(Contract).options(
        joinedload(Contract.end_customer),
        joinedload(Contract.owner),
    ).filter(Contract.id == contract_id).first()
    if not contract:
        raise NotFoundError("사업을 찾을 수 없습니다.")

    if period_year:
        year_filter = f"{period_year}-%"
        years = [period_year]
    else:
        year_rows = (
            db.query(func.distinct(func.substr(TransactionLine.revenue_month, 1, 4)))
            .filter(TransactionLine.contract_id == contract_id)
            .all()
        )
        rcpt_years = (
            db.query(func.distinct(func.substr(Receipt.revenue_month, 1, 4)))
            .filter(Receipt.contract_id == contract_id)
            .all()
        )
        years = sorted({r[0] for r in year_rows} | {r[0] for r in rcpt_years if r[0]})
        if not years:
            years = [datetime.date.today().year]
        year_filter = None

    act_q = (
        db.query(TransactionLine)
        .options(joinedload(TransactionLine.customer))
        .filter(TransactionLine.contract_id == contract_id, TransactionLine.status == STATUS_CONFIRMED)
    )
    if year_filter:
        act_q = act_q.filter(TransactionLine.revenue_month.like(year_filter))
    transaction_lines = act_q.order_by(TransactionLine.revenue_month).all()

    rcpt_q = db.query(Receipt).options(joinedload(Receipt.customer)).filter(Receipt.contract_id == contract_id)
    if year_filter:
        rcpt_q = rcpt_q.filter(Receipt.revenue_month.like(year_filter))
    receipts = rcpt_q.order_by(Receipt.receipt_date).all()

    from app.models.contract_contact import ContractContact
    from app.models.contract_period import ContractPeriod as _DP
    from sqlalchemy.orm import joinedload as _jl
    period_ids = [p.id for p in db.query(_DP.id).filter(_DP.contract_id == contract_id).all()]
    contacts = (
        db.query(ContractContact)
        .options(_jl(ContractContact.customer_contact))
        .filter(ContractContact.contract_period_id.in_(period_ids))
        .all()
    ) if period_ids else []
    contact_map: dict[int, dict] = {}
    for c in contacts:
        cc = c.customer_contact
        if c.customer_id and c.customer_id not in contact_map and cc:
            contact_map[c.customer_id] = {
                "name": cc.name, "phone": cc.phone,
                "email": cc.email, "contact_type": c.contact_type,
            }

    active_months: set[str] = set()
    for a in transaction_lines:
        active_months.add(a.revenue_month)
    for p in receipts:
        if p.revenue_month:
            active_months.add(p.revenue_month)
    sorted_months = sorted(active_months)

    revenue_by_customer: dict[int | None, dict[str, int]] = {}
    cost_by_customer: dict[int | None, dict[str, int]] = {}

    for a in transaction_lines:
        cid = a.customer_id
        target = revenue_by_customer if a.line_type == "revenue" else cost_by_customer
        row = target.setdefault(cid, {})
        row[a.revenue_month] = row.get(a.revenue_month, 0) + (a.supply_amount or 0)

    receipt_by_customer: dict[int | None, dict[str, int]] = {}
    for p in receipts:
        cid = p.customer_id
        row = receipt_by_customer.setdefault(cid, {})
        month = p.revenue_month or "unknown"
        row[month] = row.get(month, 0) + (p.amount or 0)

    customer_names: dict[int | None, str] = {None: "(미지정)"}
    for a in transaction_lines:
        if a.customer_id and a.customer_id not in customer_names:
            customer_names[a.customer_id] = a.customer.name if a.customer else f"ID:{a.customer_id}"
    for p in receipts:
        if p.customer_id and p.customer_id not in customer_names:
            customer_names[p.customer_id] = p.customer.name if p.customer else f"ID:{p.customer_id}"

    def _build_rows(by_customer: dict, section: str) -> list[dict]:
        rows = []
        for cid, month_data in sorted(by_customer.items(), key=lambda x: customer_names.get(x[0], "")):
            contact = contact_map.get(cid, {})
            rows.append({
                "section": section,
                "customer_id": cid,
                "customer_name": customer_names.get(cid, "(미지정)"),
                "contact_name": contact.get("name"),
                "contact_phone": contact.get("phone"),
                "contact_email": contact.get("email"),
                "months": month_data,
                "total": sum(month_data.values()),
            })
        return rows

    revenue_rows = _build_rows(revenue_by_customer, "revenue")
    cost_rows = _build_rows(cost_by_customer, "cost")
    receipt_rows = _build_rows(receipt_by_customer, "receipt")

    revenue_totals: dict[str, int] = {}
    cost_totals: dict[str, int] = {}
    receipt_totals: dict[str, int] = {}
    matched_totals: dict[str, int] = {}

    for m in sorted_months:
        revenue_totals[m] = sum(r["months"].get(m, 0) for r in revenue_rows)
        cost_totals[m] = sum(r["months"].get(m, 0) for r in cost_rows)
        receipt_totals[m] = sum(r["months"].get(m, 0) for r in receipt_rows)

    matched_rows = (
        db.query(
            TransactionLine.revenue_month,
            func.sum(ReceiptMatch.matched_amount),
        )
        .join(ReceiptMatch, ReceiptMatch.transaction_line_id == TransactionLine.id)
        .join(Receipt, Receipt.id == ReceiptMatch.receipt_id)
        .filter(
            TransactionLine.contract_id == contract_id,
            Receipt.contract_id == contract_id,
            TransactionLine.status == STATUS_CONFIRMED,
        )
    )
    if year_filter:
        matched_rows = matched_rows.filter(TransactionLine.revenue_month.like(year_filter))
    for month, total in matched_rows.group_by(TransactionLine.revenue_month).all():
        matched_totals[month] = int(total or 0)
        if month not in revenue_totals and month not in cost_totals and month not in receipt_totals:
            sorted_months.append(month)

    sorted_months = sorted(set(sorted_months))

    gp_monthly: dict[str, int] = {}
    gp_pct_monthly: dict[str, float | None] = {}
    ar_monthly: dict[str, int] = {}
    cumulative_revenue = 0
    cumulative_matched = 0

    for m in sorted_months:
        s = revenue_totals.get(m, 0)
        c = cost_totals.get(m, 0)
        matched = matched_totals.get(m, 0)
        gp_monthly[m] = s - c
        gp_pct_monthly[m] = round((s - c) / s * 100, 1) if s > 0 else None
        cumulative_revenue += s
        cumulative_matched += matched
        ar_monthly[m] = cumulative_revenue - cumulative_matched

    grand_revenue = sum(revenue_totals.values())
    grand_cost = sum(cost_totals.values())
    grand_receipt = sum(receipt_totals.values())
    grand_matched = sum(matched_totals.values())

    return {
        "contract_id": contract.id,
        "contract_name": contract.contract_name,
        "contract_type": contract.contract_type,
        "end_customer_name": contract.end_customer.name if contract.end_customer else None,
        "owner_name": contract.owner.name if contract.owner else None,
        "years": [int(y) for y in years],
        "months": sorted_months,
        "revenue_rows": revenue_rows,
        "revenue_totals": revenue_totals,
        "cost_rows": cost_rows,
        "cost_totals": cost_totals,
        "gp_monthly": gp_monthly,
        "gp_pct_monthly": gp_pct_monthly,
        "receipt_rows": receipt_rows,
        "receipt_totals": receipt_totals,
        "ar_monthly": ar_monthly,
        "grand_revenue": grand_revenue,
        "grand_cost": grand_cost,
        "grand_gp": grand_revenue - grand_cost,
        "grand_gp_pct": round((grand_revenue - grand_cost) / grand_revenue * 100, 1) if grand_revenue > 0 else None,
        "grand_receipt": grand_receipt,
        "grand_ar": grand_revenue - grand_matched,
    }


# ═══ Excel Export ════════════════════════════════════════════════


def export_summary(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """요약 현황 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = get_summary(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "요약 현황"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="요약 현황").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    # 월별 테이블
    row = 4
    headers = ["월", "Forecast 매출", "Actual 매출", "매입", "GP", "GP%", "입금", "미수금"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, mr in enumerate(data["period_summary"], row + 1):
        vals = [mr["month"], mr["forecast_revenue"], mr["actual_revenue"], mr["cost"],
                mr["gp"], mr["gp_pct"], mr["receipt"], mr["ar"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci >= 2 and ci != 6:
                cell.number_format = number_fmt

    # 합계 행
    kpi = data["kpis"]
    sum_row = len(data["period_summary"]) + row + 1
    sum_vals = ["합계", kpi["forecast_revenue"], kpi["actual_revenue"], kpi["cost"],
                kpi["gp"], kpi["gp_pct"], kpi["receipt"], kpi["ar"]]
    for ci, v in enumerate(sum_vals, 1):
        cell = ws.cell(row=sum_row, column=ci, value=v)
        cell.font = header_font
        cell.border = thin_border
        if ci >= 2 and ci != 6:
            cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 10
    for ci in range(2, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 15

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_forecast_vs_actual(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """Forecast vs Actual Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = list_forecast_vs_actual(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast vs Actual"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Forecast vs Actual").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    headers = ["사업명", "사업유형", "담당", "부서", "END고객", "단계",
               "Forecast", "Actual", "Gap", "달성률(%)", "GP", "GP%"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, row_data in enumerate(data["rows"], 5):
        vals = [
            row_data["contract_name"], row_data["contract_type"], row_data["owner_name"],
            row_data["department"], row_data["end_customer_name"], row_data["stage"],
            row_data["forecast_revenue"], row_data["actual_revenue"],
            row_data["gap_revenue"], row_data["achievement_rate"],
            row_data["gp"], row_data["gp_pct"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci in (7, 8, 9, 11):
                cell.number_format = number_fmt

    if data["totals"]:
        t = data["totals"]
        ri = len(data["rows"]) + 5
        ws.cell(row=ri, column=1, value="합계").font = header_font
        for ci, v in [(7, t["forecast_revenue"]), (8, t["actual_revenue"]),
                       (9, t["gap_revenue"]), (10, t["achievement_rate"]),
                       (11, t["gp"]), (12, t["gp_pct"])]:
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = header_font
            cell.border = thin_border
            if ci in (7, 8, 9, 11):
                cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    for ci in range(7, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_receivables(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """미수 현황 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = list_receivables(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미수 현황"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ar_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="미수 현황").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    headers = ["사업명", "사업유형", "담당", "부서", "END고객", "매출 확정", "입금", "미수금", "미수율(%)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, row_data in enumerate(data["rows"], 5):
        vals = [
            row_data["contract_name"], row_data["contract_type"], row_data["owner_name"],
            row_data["department"], row_data["end_customer_name"],
            row_data["actual_revenue"], row_data["receipt"],
            row_data["ar"], row_data["ar_rate"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci in (6, 7, 8):
                cell.number_format = number_fmt
            if ci == 8 and isinstance(v, (int, float)) and v > 0:
                cell.font = Font(color="CC0000", bold=True)
                cell.fill = ar_fill

    t = data["totals"]
    ri = len(data["rows"]) + 5
    ws.cell(row=ri, column=1, value="합계").font = header_font
    for ci, v in [(6, t["actual_revenue"]), (7, t["receipt"]), (8, t["ar"]), (9, t["ar_rate"])]:
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = header_font
        cell.border = thin_border
        if ci in (6, 7, 8):
            cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    for ci in range(6, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_contract_pnl(
    db: Session,
    contract_id: int,
    period_year: int | None = None,
) -> bytes:
    """매입매출관리 보고서 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = get_contract_pnl(db, contract_id, period_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "매입매출관리"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ar_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    months = data["months"]

    ws.cell(row=1, column=2, value=f"▣ {data['contract_name']}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=2, value="[단위:원,VAT별도]").font = Font(size=9, color="666666")

    row = 4
    headers = ["", "거래처명", "담당자", "연락처", "이메일"]
    for m in months:
        headers.append(m[:7])
    headers.append("합계")
    for c_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    def _write_section(start_row: int, section_name: str, rows_data: list[dict], totals: dict, fill: PatternFill) -> int:
        r = start_row
        for item in rows_data:
            ws.cell(row=r, column=2, value=item["customer_name"]).border = thin_border
            ws.cell(row=r, column=3, value=item.get("contact_name")).border = thin_border
            ws.cell(row=r, column=4, value=item.get("contact_phone")).border = thin_border
            ws.cell(row=r, column=5, value=item.get("contact_email")).border = thin_border
            for mi, m in enumerate(months):
                val = item["months"].get(m, 0)
                cell = ws.cell(row=r, column=6 + mi, value=val if val else None)
                cell.number_format = number_fmt
                cell.border = thin_border
            total_cell = ws.cell(row=r, column=6 + len(months), value=item["total"])
            total_cell.number_format = number_fmt
            total_cell.border = thin_border
            r += 1

        label = {"revenue": "[매출]", "cost": "[매입]", "receipt": "[입금]"}.get(section_name, "")
        ws.cell(row=r, column=2, value=f"{label} 합계").font = header_font
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = thin_border
        for mi, m in enumerate(months):
            val = totals.get(m, 0)
            cell = ws.cell(row=r, column=6 + mi, value=val if val else None)
            cell.number_format = number_fmt
            cell.font = header_font
            cell.fill = fill
            cell.border = thin_border
        grand = sum(totals.values())
        total_cell = ws.cell(row=r, column=6 + len(months), value=grand)
        total_cell.number_format = number_fmt
        total_cell.font = header_font
        total_cell.fill = fill
        total_cell.border = thin_border
        return r + 2

    cur_row = 5
    if data["revenue_rows"]:
        cur_row = _write_section(cur_row, "revenue", data["revenue_rows"], data["revenue_totals"], section_fill)
    else:
        ws.cell(row=cur_row, column=2, value="[매출] 합계").font = header_font
        cur_row += 2

    if data["cost_rows"]:
        cur_row = _write_section(cur_row, "cost", data["cost_rows"], data["cost_totals"], section_fill)
    else:
        ws.cell(row=cur_row, column=2, value="[매입] 합계").font = header_font
        cur_row += 2

    ws.cell(row=cur_row, column=2, value="GP").font = header_font
    ws.cell(row=cur_row, column=2).fill = gp_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["gp_monthly"].get(m, 0)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val)
        cell.number_format = number_fmt
        cell.font = header_font
        cell.fill = gp_fill
        cell.border = thin_border
    ws.cell(row=cur_row, column=6 + len(months), value=data["grand_gp"]).number_format = number_fmt
    ws.cell(row=cur_row, column=6 + len(months)).font = header_font
    ws.cell(row=cur_row, column=6 + len(months)).fill = gp_fill
    ws.cell(row=cur_row, column=6 + len(months)).border = thin_border
    cur_row += 1

    ws.cell(row=cur_row, column=2, value="GP%").font = header_font
    ws.cell(row=cur_row, column=2).fill = gp_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["gp_pct_monthly"].get(m)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val / 100 if val is not None else None)
        cell.number_format = '0.0%'
        cell.fill = gp_fill
        cell.border = thin_border
    if data["grand_gp_pct"] is not None:
        cell = ws.cell(row=cur_row, column=6 + len(months), value=data["grand_gp_pct"] / 100)
        cell.number_format = '0.0%'
        cell.fill = gp_fill
        cell.border = thin_border
    cur_row += 2

    if data["receipt_rows"]:
        cur_row = _write_section(cur_row, "receipt", data["receipt_rows"], data["receipt_totals"], section_fill)

    ws.cell(row=cur_row, column=2, value="미수금").font = Font(bold=True, size=10, color="CC0000")
    ws.cell(row=cur_row, column=2).fill = ar_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["ar_monthly"].get(m, 0)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val)
        cell.number_format = number_fmt
        cell.fill = ar_fill
        cell.border = thin_border
        if val > 0:
            cell.font = Font(color="CC0000")
    ws.cell(row=cur_row, column=6 + len(months), value=data["grand_ar"]).number_format = number_fmt
    ws.cell(row=cur_row, column=6 + len(months)).fill = ar_fill
    ws.cell(row=cur_row, column=6 + len(months)).border = thin_border

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    for mi in range(len(months) + 1):
        col_letter = openpyxl.utils.get_column_letter(6 + mi)
        ws.column_dimensions[col_letter].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
