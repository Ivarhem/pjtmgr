"""Excel 템플릿 생성 서비스"""
from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def _header_style(ws: "Worksheet", row: int, cols: list[str], fill_color: str = "1a3a5c") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def build_template_contracts(contract_type_codes: list[str] | None = None) -> bytes:
    """Sheet1(영업기회) 전용 블랭크 템플릿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "영업기회"
    cols = ["연도", "번호", "사업유형", "담당", "거래처(END)", "영업기회명", "진행단계", "예상매출(원)", "예상GP(원)"]
    _header_style(ws, 1, cols)
    ws.row_dimensions[1].height = 20
    for i, w in enumerate([8, 8, 10, 12, 20, 50, 12, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    types_str = "/".join(contract_type_codes) if contract_type_codes else "MA/SI/HW/TS/Prod/ETC"
    note = ws.cell(row=2, column=1, value=f"[사업유형: {types_str}  진행단계: 10%/50%/70%/90%/계약완료]")
    note.font = Font(color="999999", italic=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_forecast(periods: list[dict]) -> bytes:
    """Sheet2(월별계획) 템플릿 — 기존 등록 영업기회 목록 포함.
    periods: [{"period_id": int, "period_year": int, "contract_name": str, "contract_code": str}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "월별계획"
    month_pairs = [f for m in range(1, 13) for f in [f"{m}월매출", f"{m}월GP"]]
    cols = ["기간ID", "연도", "사업명(참고)"] + month_pairs
    _header_style(ws, 1, cols, fill_color="254d7a")
    ws.row_dimensions[1].height = 20
    for i, w in enumerate([10, 8, 40] + [12] * 24, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 참고용 안내
    note_font = Font(color="999999", italic=True, size=9)

    # 기존 영업기회 목록 pre-fill (금액 칸은 비워둠)
    ref_fill = PatternFill("solid", fgColor="f0f4fa")
    for row_idx, p in enumerate(periods, start=2):
        ws.cell(row=row_idx, column=1, value=p["period_id"])
        ws.cell(row=row_idx, column=2, value=p["period_year"])
        c = ws.cell(row=row_idx, column=3, value=p["contract_name"])
        c.font = note_font
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).fill = ref_fill

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_actuals(periods: list[dict]) -> bytes:
    """Sheet3(실적) 템플릿 — 기존 등록 영업기회 목록 포함.
    periods: [{"period_id": int, "period_year": int, "contract_name": str}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "실적"
    month_cols = [f"{m}월" for m in range(1, 13)]
    cols = ["기간ID", "연도", "사업명(참고)", "매출/매입", "거래처명", "세금계산서담당자", "연락처", "이메일"] + month_cols
    _header_style(ws, 1, cols, fill_color="2a7a3b")
    ws.row_dimensions[1].height = 20
    for i, w in enumerate([10, 8, 40, 10, 20, 15, 15, 25] + [12] * 12, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note_font = Font(color="999999", italic=True, size=9)
    ref_fill = PatternFill("solid", fgColor="f0f9f2")
    for row_idx, p in enumerate(periods, start=2):
        ws.cell(row=row_idx, column=1, value=p["period_id"])
        ws.cell(row=row_idx, column=2, value=p["period_year"])
        c = ws.cell(row=row_idx, column=3, value=p["contract_name"])
        c.font = note_font
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).fill = ref_fill

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template(contract_type_codes: list[str] | None = None) -> bytes:
    wb = Workbook()

    # ── Sheet1: 영업기회 ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "영업기회"
    cols1 = ["연도", "번호", "사업유형", "담당", "거래처(END)", "영업기회명", "진행단계", "예상매출(원)", "예상GP(원)"]
    _header_style(ws1, 1, cols1)
    ws1.row_dimensions[1].height = 20
    # 컬럼 너비
    widths1 = [8, 8, 10, 12, 20, 50, 12, 16, 16]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    # 안내 행
    types_str = "/".join(contract_type_codes) if contract_type_codes else "MA/SI/HW/TS/Prod/ETC"
    note1 = ws1.cell(row=2, column=1, value=f"[사업유형: {types_str}  진행단계: 10%/50%/70%/90%/계약완료]")
    note1.font = Font(color="999999", italic=True)

    # ── Sheet2: 월별계획 ──────────────────────────────────────────
    ws2 = wb.create_sheet("월별계획")
    month_pairs = []
    for m in range(1, 13):
        month_pairs += [f"{m}월매출", f"{m}월GP"]
    cols2 = ["연도", "번호"] + month_pairs
    _header_style(ws2, 1, cols2, fill_color="254d7a")
    ws2.row_dimensions[1].height = 20
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 8
    for i in range(3, len(cols2) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 12
    note2 = ws2.cell(row=2, column=1, value="[단위: 원, VAT별도  |  연도+번호는 Sheet1과 동일하게]")
    note2.font = Font(color="999999", italic=True)

    # ── Sheet3: 실적 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("실적")
    fixed3 = ["연도", "번호", "매출/매입", "거래처명", "세금계산서담당자", "연락처", "이메일"]
    month_cols = [f"{m}월" for m in range(1, 13)]
    cols3 = fixed3 + month_cols
    _header_style(ws3, 1, cols3, fill_color="2a7a3b")
    ws3.row_dimensions[1].height = 20
    widths3 = [8, 8, 10, 20, 15, 15, 25] + [12] * 12
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    note3 = ws3.cell(row=2, column=1, value="[매출/매입: '매출' 또는 '매입'  |  단위: 원, VAT별도  |  귀속월 기준 공급가액]")
    note3.font = Font(color="999999", italic=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
