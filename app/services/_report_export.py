"""보고서 Excel Export 함수.

report.py에서 분리된 Excel 다운로드 전용 모듈.
"""
from __future__ import annotations

from typing import TYPE_CHECKING

from sqlalchemy.orm import Session

from app.schemas.report import ReportFilter
from app.services.report import (
    get_summary,
    list_forecast_vs_actual,
    list_receivables,
    get_contract_pnl,
)

if TYPE_CHECKING:
    from app.models.user import User


def export_summary(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """요약 현황 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = get_summary(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "요약 현황"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="요약 현황").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    # 월별 테이블
    row = 4
    headers = ["월", "Forecast 매출", "Actual 매출", "매입", "GP", "GP%", "입금", "미수금"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, mr in enumerate(data["period_summary"], row + 1):
        vals = [mr["month"], mr["forecast_revenue"], mr["actual_revenue"], mr["cost"],
                mr["gp"], mr["gp_pct"], mr["receipt"], mr["ar"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci >= 2 and ci != 6:
                cell.number_format = number_fmt

    # 합계 행
    kpi = data["kpis"]
    sum_row = len(data["period_summary"]) + row + 1
    sum_vals = ["합계", kpi["forecast_revenue"], kpi["actual_revenue"], kpi["cost"],
                kpi["gp"], kpi["gp_pct"], kpi["receipt"], kpi["ar"]]
    for ci, v in enumerate(sum_vals, 1):
        cell = ws.cell(row=sum_row, column=ci, value=v)
        cell.font = header_font
        cell.border = thin_border
        if ci >= 2 and ci != 6:
            cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 10
    for ci in range(2, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 15

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_forecast_vs_actual(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """Forecast vs Actual Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = list_forecast_vs_actual(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast vs Actual"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Forecast vs Actual").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    headers = ["사업명", "사업유형", "담당", "부서", "END고객", "단계",
               "Forecast", "Actual", "Gap", "달성률(%)", "GP", "GP%"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, row_data in enumerate(data["rows"], 5):
        vals = [
            row_data["contract_name"], row_data["contract_type"], row_data["owner_name"],
            row_data["department"], row_data["end_customer_name"], row_data["stage"],
            row_data["forecast_revenue"], row_data["actual_revenue"],
            row_data["gap_revenue"], row_data["achievement_rate"],
            row_data["gp"], row_data["gp_pct"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci in (7, 8, 9, 11):
                cell.number_format = number_fmt

    if data["totals"]:
        t = data["totals"]
        ri = len(data["rows"]) + 5
        ws.cell(row=ri, column=1, value="합계").font = header_font
        for ci, v in [(7, t["forecast_revenue"]), (8, t["actual_revenue"]),
                       (9, t["gap_revenue"]), (10, t["achievement_rate"]),
                       (11, t["gp"]), (12, t["gp_pct"])]:
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = header_font
            cell.border = thin_border
            if ci in (7, 8, 9, 11):
                cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    for ci in range(7, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_receivables(
    db: Session,
    filt: ReportFilter,
    *,
    current_user: User | None = None,
) -> bytes:
    """미수 현황 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = list_receivables(db, filt, current_user=current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미수 현황"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ar_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="미수 현황").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"기간: {filt.date_from[:7]} ~ {filt.date_to[:7]}").font = Font(size=10, color="666666")

    headers = ["사업명", "사업유형", "담당", "부서", "END고객", "매출 확정", "입금", "미수금", "미수율(%)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for ri, row_data in enumerate(data["rows"], 5):
        vals = [
            row_data["contract_name"], row_data["contract_type"], row_data["owner_name"],
            row_data["department"], row_data["end_customer_name"],
            row_data["actual_revenue"], row_data["receipt"],
            row_data["ar"], row_data["ar_rate"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci in (6, 7, 8):
                cell.number_format = number_fmt
            if ci == 8 and isinstance(v, (int, float)) and v > 0:
                cell.font = Font(color="CC0000", bold=True)
                cell.fill = ar_fill

    t = data["totals"]
    ri = len(data["rows"]) + 5
    ws.cell(row=ri, column=1, value="합계").font = header_font
    for ci, v in [(6, t["actual_revenue"]), (7, t["receipt"]), (8, t["ar"]), (9, t["ar_rate"])]:
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = header_font
        cell.border = thin_border
        if ci in (6, 7, 8):
            cell.number_format = number_fmt

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    for ci in range(6, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_contract_pnl(
    db: Session,
    contract_id: int,
    period_year: int | None = None,
) -> bytes:
    """매입매출관리 보고서 Excel 다운로드."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = get_contract_pnl(db, contract_id, period_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "매입매출관리"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ar_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    months = data["months"]

    ws.cell(row=1, column=2, value=f"▣ {data['contract_name']}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=2, value="[단위:원,VAT별도]").font = Font(size=9, color="666666")

    row = 4
    headers = ["", "거래처명", "담당자", "연락처", "이메일"]
    for m in months:
        headers.append(m[:7])
    headers.append("합계")
    for c_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    def _write_section(start_row: int, section_name: str, rows_data: list[dict], totals: dict, fill: PatternFill) -> int:
        r = start_row
        for item in rows_data:
            ws.cell(row=r, column=2, value=item["customer_name"]).border = thin_border
            ws.cell(row=r, column=3, value=item.get("contact_name")).border = thin_border
            ws.cell(row=r, column=4, value=item.get("contact_phone")).border = thin_border
            ws.cell(row=r, column=5, value=item.get("contact_email")).border = thin_border
            for mi, m in enumerate(months):
                val = item["months"].get(m, 0)
                cell = ws.cell(row=r, column=6 + mi, value=val if val else None)
                cell.number_format = number_fmt
                cell.border = thin_border
            total_cell = ws.cell(row=r, column=6 + len(months), value=item["total"])
            total_cell.number_format = number_fmt
            total_cell.border = thin_border
            r += 1

        label = {"revenue": "[매출]", "cost": "[매입]", "receipt": "[입금]"}.get(section_name, "")
        ws.cell(row=r, column=2, value=f"{label} 합계").font = header_font
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = thin_border
        for mi, m in enumerate(months):
            val = totals.get(m, 0)
            cell = ws.cell(row=r, column=6 + mi, value=val if val else None)
            cell.number_format = number_fmt
            cell.font = header_font
            cell.fill = fill
            cell.border = thin_border
        grand = sum(totals.values())
        total_cell = ws.cell(row=r, column=6 + len(months), value=grand)
        total_cell.number_format = number_fmt
        total_cell.font = header_font
        total_cell.fill = fill
        total_cell.border = thin_border
        return r + 2

    cur_row = 5
    if data["revenue_rows"]:
        cur_row = _write_section(cur_row, "revenue", data["revenue_rows"], data["revenue_totals"], section_fill)
    else:
        ws.cell(row=cur_row, column=2, value="[매출] 합계").font = header_font
        cur_row += 2

    if data["cost_rows"]:
        cur_row = _write_section(cur_row, "cost", data["cost_rows"], data["cost_totals"], section_fill)
    else:
        ws.cell(row=cur_row, column=2, value="[매입] 합계").font = header_font
        cur_row += 2

    ws.cell(row=cur_row, column=2, value="GP").font = header_font
    ws.cell(row=cur_row, column=2).fill = gp_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["gp_monthly"].get(m, 0)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val)
        cell.number_format = number_fmt
        cell.font = header_font
        cell.fill = gp_fill
        cell.border = thin_border
    ws.cell(row=cur_row, column=6 + len(months), value=data["grand_gp"]).number_format = number_fmt
    ws.cell(row=cur_row, column=6 + len(months)).font = header_font
    ws.cell(row=cur_row, column=6 + len(months)).fill = gp_fill
    ws.cell(row=cur_row, column=6 + len(months)).border = thin_border
    cur_row += 1

    ws.cell(row=cur_row, column=2, value="GP%").font = header_font
    ws.cell(row=cur_row, column=2).fill = gp_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["gp_pct_monthly"].get(m)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val / 100 if val is not None else None)
        cell.number_format = '0.0%'
        cell.fill = gp_fill
        cell.border = thin_border
    if data["grand_gp_pct"] is not None:
        cell = ws.cell(row=cur_row, column=6 + len(months), value=data["grand_gp_pct"] / 100)
        cell.number_format = '0.0%'
        cell.fill = gp_fill
        cell.border = thin_border
    cur_row += 2

    if data["receipt_rows"]:
        cur_row = _write_section(cur_row, "receipt", data["receipt_rows"], data["receipt_totals"], section_fill)

    ws.cell(row=cur_row, column=2, value="미수금").font = Font(bold=True, size=10, color="CC0000")
    ws.cell(row=cur_row, column=2).fill = ar_fill
    ws.cell(row=cur_row, column=2).border = thin_border
    for mi, m in enumerate(months):
        val = data["ar_monthly"].get(m, 0)
        cell = ws.cell(row=cur_row, column=6 + mi, value=val)
        cell.number_format = number_fmt
        cell.fill = ar_fill
        cell.border = thin_border
        if val > 0:
            cell.font = Font(color="CC0000")
    ws.cell(row=cur_row, column=6 + len(months), value=data["grand_ar"]).number_format = number_fmt
    ws.cell(row=cur_row, column=6 + len(months)).fill = ar_fill
    ws.cell(row=cur_row, column=6 + len(months)).border = thin_border

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    for mi in range(len(months) + 1):
        col_letter = openpyxl.utils.get_column_letter(6 + mi)
        ws.column_dimensions[col_letter].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
