"""인프라모듈 Excel Export 서비스 — 업체 단위 데이터 내보내기."""
from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import NotFoundError
from app.modules.common.models.partner import Partner
from app.modules.infra.models.asset import Asset
from app.modules.infra.models.asset_interface import AssetInterface
from app.modules.infra.models.ip_subnet import IpSubnet
from app.modules.infra.models.port_map import PortMap
from app.modules.infra.models.period_asset import PeriodAsset
from app.modules.infra.services.network_service import build_interface_map

# -- 헤더 / 필드 매핑 --

_INVENTORY_HEADERS = [
    ("Seq", None),
    ("자산명", "asset_name"),
    ("상태", "status"),
    ("환경", "environment"),
    ("위치", "location"),
    ("센터", "center"),
    ("운영유형", "operation_type"),
    ("장비ID", "equipment_id"),
    ("랙 No", "rack_no"),
    ("랙 Unit", "rack_unit"),
    ("Phase", "phase"),
    ("입고일", "received_date"),
    ("분류", "category"),
    ("세분류", "subcategory"),
    ("제조사", "vendor"),
    ("모델", "model"),
    ("S/N", "serial_no"),
    ("호스트명", "hostname"),
    ("클러스터", "cluster"),
    ("서비스명", "service_name"),
    ("Zone", "zone"),
    ("Size(U)", "size_unit"),
    ("LC", "lc_count"),
    ("HA", "ha_count"),
    ("UTP", "utp_count"),
    ("전원 수", "power_count"),
    ("전원유형", "power_type"),
    ("FW버전", "firmware_version"),
    ("자산분류", "asset_class"),
    ("자산번호", "asset_number"),
    ("취득년도", "year_acquired"),
    ("부서", "dept"),
    ("주담당", "primary_contact_name"),
    ("부담당", "secondary_contact_name"),
    ("유지보수사", "maintenance_vendor"),
    ("비고", "note"),
]

_SUBNET_HEADERS = [
    ("대역명", "name"),
    ("Subnet (CIDR)", "subnet"),
    ("Netmask", "netmask"),
    ("Gateway", "gateway"),
    ("VLAN", "vlan_id"),
    ("Zone", "zone"),
    ("역할", "role"),
    ("지역/층", "region"),
    ("분류", "category"),
    ("설명", "description"),
]

_PORTMAP_HEADERS = [
    ("Seq", "seq"),
    ("요약", "summary"),
    ("연결유형", "connection_type"),
    ("출발 자산", None),        # denormalized via interface map
    ("출발 IF", None),          # denormalized via interface map
    ("출발 호스트", None),      # denormalized via interface map
    ("출발 Zone", None),        # denormalized via interface map
    ("도착 자산", None),        # denormalized via interface map
    ("도착 IF", None),          # denormalized via interface map
    ("도착 호스트", None),      # denormalized via interface map
    ("도착 Zone", None),        # denormalized via interface map
    ("Protocol", "protocol"),
    ("Port", "port"),
    ("용도", "purpose"),
    ("상태", "status"),
    ("케이블유형", "cable_type"),
    ("케이블속도", "cable_speed"),
    ("케이블번호", "cable_no"),
    ("비고", "note"),
]

_INTERFACE_HEADERS = [
    ("자산명", None),           # denormalized from Asset
    ("IF이름", "name"),
    ("유형", "if_type"),
    ("속도", "speed"),
    ("미디어", "media_type"),
    ("슬롯", "slot"),
    ("Admin", "admin_status"),
    ("Oper", "oper_status"),
    ("MAC", "mac_address"),
    ("설명", "description"),
]

# -- 스타일 --

_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_header(ws, headers: list[tuple[str, str | None]]) -> None:
    for col_idx, (label, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws, headers: list[tuple[str, str | None]]) -> None:
    for col_idx, (label, _) in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(label) + 2
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)) + 1)
        ws.column_dimensions[col_letter].width = min(max_len + 1, 40)


# -- 메인 --


def export_partner(db: Session, partner_id: int, period_id: int | None = None) -> bytes:
    """업체 데이터를 4개 시트(Inventory, IP대역, Portmap, Interfaces)로 Export. period_id 지정 시 해당 기간 자산만 필터."""
    if db.get(Partner, partner_id) is None:
        raise NotFoundError("Partner not found")

    # 기간 필터용 자산 ID 집합
    period_asset_ids: set[int] | None = None
    if period_id:
        period_asset_ids = set(db.scalars(
            select(PeriodAsset.asset_id).where(PeriodAsset.contract_period_id == period_id)
        ))

    wb = openpyxl.Workbook()

    # Sheet 1: Inventory (Assets)
    ws_inv = wb.active
    ws_inv.title = "01. Inventory"
    stmt = select(Asset).where(Asset.partner_id == partner_id).order_by(Asset.id.asc())
    if period_asset_ids is not None:
        stmt = stmt.where(Asset.id.in_(period_asset_ids)) if period_asset_ids else stmt.where(Asset.id == -1)
    assets = list(db.scalars(stmt))
    _write_header(ws_inv, _INVENTORY_HEADERS)
    for row_idx, asset in enumerate(assets, 2):
        for col_idx, (_, field) in enumerate(_INVENTORY_HEADERS, 1):
            if field is None:
                ws_inv.cell(row=row_idx, column=col_idx, value=row_idx - 1)
            else:
                val = getattr(asset, field, None)
                if val is not None:
                    val = str(val)
                ws_inv.cell(row=row_idx, column=col_idx, value=val)
    _auto_width(ws_inv, _INVENTORY_HEADERS)

    # Sheet 2: IP 대역
    ws_sub = wb.create_sheet("05. 네트워크 대역")
    subnets = list(
        db.scalars(
            select(IpSubnet)
            .where(IpSubnet.partner_id == partner_id)
            .order_by(IpSubnet.id.asc())
        )
    )
    _write_header(ws_sub, _SUBNET_HEADERS)
    for row_idx, sub in enumerate(subnets, 2):
        for col_idx, (_, field) in enumerate(_SUBNET_HEADERS, 1):
            val = getattr(sub, field, None)
            if val is not None:
                val = str(val)
            ws_sub.cell(row=row_idx, column=col_idx, value=val)
    _auto_width(ws_sub, _SUBNET_HEADERS)

    # Sheet 3: Portmap
    ws_pm = wb.create_sheet("03. Portmap")
    portmaps = list(
        db.scalars(
            select(PortMap)
            .where(PortMap.partner_id == partner_id)
            .order_by(PortMap.id.asc())
        )
    )
    iface_map = build_interface_map(db, portmaps)
    _write_header(ws_pm, _PORTMAP_HEADERS)
    for row_idx, pm in enumerate(portmaps, 2):
        src = iface_map.get(pm.src_interface_id) if pm.src_interface_id is not None else None
        dst = iface_map.get(pm.dst_interface_id) if pm.dst_interface_id is not None else None
        # Denormalized values keyed by header label
        denorm: dict[str, object] = {
            "출발 자산": src["asset_name"] if src else None,
            "출발 IF": src["iface_name"] if src else None,
            "출발 호스트": src["hostname"] if src else None,
            "출발 Zone": src["zone"] if src else None,
            "도착 자산": dst["asset_name"] if dst else None,
            "도착 IF": dst["iface_name"] if dst else None,
            "도착 호스트": dst["hostname"] if dst else None,
            "도착 Zone": dst["zone"] if dst else None,
        }
        for col_idx, (label, field) in enumerate(_PORTMAP_HEADERS, 1):
            if field is None:
                val = denorm.get(label)
            else:
                val = getattr(pm, field, None)
            if val is not None:
                val = str(val) if not isinstance(val, int) else val
            ws_pm.cell(row=row_idx, column=col_idx, value=val)
    _auto_width(ws_pm, _PORTMAP_HEADERS)

    # Sheet 4: Interfaces
    ws_if = wb.create_sheet("04. Interfaces")
    iface_stmt = (
        select(AssetInterface, Asset.asset_name)
        .join(Asset, AssetInterface.asset_id == Asset.id)
        .where(Asset.partner_id == partner_id)
        .order_by(Asset.asset_name, AssetInterface.sort_order, AssetInterface.id)
    )
    if period_asset_ids is not None:
        iface_stmt = (
            iface_stmt.where(Asset.id.in_(period_asset_ids))
            if period_asset_ids
            else iface_stmt.where(Asset.id == -1)
        )
    ifaces = list(db.execute(iface_stmt))
    _write_header(ws_if, _INTERFACE_HEADERS)
    for row_idx, (iface, asset_name) in enumerate(ifaces, 2):
        ws_if.cell(row=row_idx, column=1, value=asset_name)  # 자산명 (denormalized)
        for col_idx, (_, field) in enumerate(_INTERFACE_HEADERS[1:], 2):
            val = getattr(iface, field, None)
            if val is not None:
                val = str(val)
            ws_if.cell(row=row_idx, column=col_idx, value=val)
    _auto_width(ws_if, _INTERFACE_HEADERS)

    # 바이트 출력
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
