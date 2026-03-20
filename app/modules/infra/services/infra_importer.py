"""인프라모듈 Excel Import 서비스 (자산/IP대역/포트맵 파싱 및 DB 저장)"""
from __future__ import annotations

from io import BytesIO

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessRuleError, NotFoundError
from app.modules.infra.models.asset import Asset
from app.modules.infra.models.ip_subnet import IpSubnet
from app.modules.infra.models.port_map import PortMap
from app.modules.common.models.customer import Customer

# ── 파일 유효성 검증 ──

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

INVENTORY_SHEET_NAME = "01. Inventory"

# Row 3 기준 컬럼 매핑 (col_index 1-based → Asset 필드)
_COLUMN_MAP: dict[int, str] = {
    # col 2: Seq → 무시 (auto)
    3: "center",
    4: "operation_type",
    5: "equipment_id",
    6: "rack_no",
    7: "rack_unit",
    8: "phase",
    9: "received_date",
    10: "category",
    11: "subcategory",
    12: "vendor",
    13: "model",
    14: "serial_no",
    15: "hostname",
    16: "cluster",
    17: "service_name",
    18: "zone",
    19: "service_ip",
    20: "mgmt_ip",
    21: "size_unit",
    22: "lc_count",
    23: "ha_count",
    24: "utp_count",
    25: "power_count",
    26: "power_type",
    27: "firmware_version",
    28: "asset_class",
    29: "asset_number",
    30: "year_acquired",
    31: "dept",
    32: "primary_contact_name",
    33: "secondary_contact_name",
    34: "maintenance_vendor",
    35: "note",
}

# 정수 변환 대상 필드
_INT_FIELDS = {"size_unit", "lc_count", "ha_count", "utp_count", "power_count", "year_acquired"}


def validate_xlsx(filename: str | None, content_type: str | None) -> None:
    """Excel 파일 확장자 및 MIME 타입 검증."""
    if not filename or not filename.lower().endswith(".xlsx"):
        raise BusinessRuleError("xlsx 파일만 업로드할 수 있습니다.", status_code=422)
    if content_type and content_type != _XLSX_MIME:
        raise BusinessRuleError("xlsx 파일만 업로드할 수 있습니다.", status_code=422)


def _to_str(val) -> str:
    """셀 값을 문자열로 변환. None → ""."""
    if val is None:
        return ""
    return str(val).strip()


def _to_int(val) -> int | None:
    """셀 값을 정수로 변환. 빈 값 → None."""
    s = _to_str(val)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _to_date(val):
    """셀 값을 date로 변환. openpyxl은 datetime으로 읽으므로 .date() 추출."""
    from datetime import date, datetime

    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = _to_str(val)
    if not s:
        return None
    # 간단한 날짜 파싱 시도
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _derive_asset_type(category: str) -> str:
    """대분류(category)에서 asset_type을 유추. 매핑 없으면 'etc'."""
    cat = category.lower() if category else ""
    mapping = {
        "서버": "server",
        "server": "server",
        "네트워크": "network",
        "network": "network",
        "스위치": "network",
        "switch": "network",
        "보안": "security",
        "security": "security",
        "방화벽": "security",
        "firewall": "security",
        "스토리지": "storage",
        "storage": "storage",
    }
    for keyword, asset_type in mapping.items():
        if keyword in cat:
            return asset_type
    return "etc"


def _append_error(
    errors: list[str],
    error_details: list[dict],
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    column: str | None = None,
    code: str | None = None,
) -> None:
    errors.append(message)
    error_details.append(
        {"message": message, "sheet": sheet, "row": row, "column": column, "code": code}
    )


def parse_inventory_sheet(file_bytes: bytes, customer_id: int) -> dict:
    """
    01. Inventory 시트를 파싱하여 프리뷰 데이터 반환.

    Returns:
        {
            "rows": [dict, ...],      # 파싱된 행 목록 (Asset 필드 딕셔너리)
            "preview_rows": [dict, ...],  # UI 프리뷰용 요약 행
            "errors": [str, ...],
            "error_details": [dict, ...],
            "warnings": [str, ...],
            "total": int,
            "valid_count": int,
        }
    """
    errors: list[str] = []
    error_details: list[dict] = []
    warnings: list[str] = []
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    # 워크북 로드
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return {
            "rows": [],
            "preview_rows": [],
            "errors": errors,
            "error_details": error_details,
            "warnings": warnings,
            "total": 0,
            "valid_count": 0,
        }

    # 시트 존재 확인
    sheet_names = wb.sheetnames
    ws = None
    for name in sheet_names:
        if name.strip() == INVENTORY_SHEET_NAME:
            ws = wb[name]
            break

    if ws is None:
        # 시트명이 정확하지 않으면 첫 번째 시트를 시도
        if sheet_names:
            warnings.append(
                f"'{INVENTORY_SHEET_NAME}' 시트를 찾지 못해 첫 번째 시트('{sheet_names[0]}')를 사용합니다."
            )
            ws = wb[sheet_names[0]]
        else:
            _append_error(errors, error_details, "빈 파일입니다.", code="empty_file")
            wb.close()
            return {
                "rows": [],
                "preview_rows": [],
                "errors": errors,
                "error_details": error_details,
                "warnings": warnings,
                "total": 0,
                "valid_count": 0,
            }

    # 데이터 행 파싱 (Row 4부터, Row 1-2는 그룹 헤더, Row 3은 컬럼 헤더)
    seen_names: dict[str, int] = {}  # asset_name → first_row for duplicate detection
    row_num = 0
    for row in ws.iter_rows(min_row=4, values_only=False):
        row_num += 1
        excel_row = row_num + 3  # 실제 엑셀 행 번호

        # 빈 행 스킵: Seq(col 2) 또는 hostname(col 15) 중 하나라도 있으면 데이터 행
        seq_val = _to_str(row[1].value) if len(row) > 1 else ""
        hostname_val = _to_str(row[14].value) if len(row) > 14 else ""
        category_val = _to_str(row[9].value) if len(row) > 9 else ""

        if not seq_val and not hostname_val and not category_val:
            continue

        # 필드 추출
        asset_data: dict = {"customer_id": None}  # filled at import time
        row_errors: list[str] = []

        for col_idx, field_name in _COLUMN_MAP.items():
            cell_val = row[col_idx - 1].value if len(row) >= col_idx else None

            if field_name == "received_date":
                asset_data[field_name] = _to_date(cell_val)
            elif field_name in _INT_FIELDS:
                asset_data[field_name] = _to_int(cell_val)
            else:
                asset_data[field_name] = _to_str(cell_val) or None

        # asset_name 결정: hostname > equipment_id > "Row {n}"
        asset_name = asset_data.get("hostname") or asset_data.get("equipment_id")
        if not asset_name:
            asset_name = f"Row-{excel_row}"
            warnings.append(f"행 {excel_row}: hostname/equipment_id 없음, '{asset_name}'으로 지정")

        asset_data["asset_name"] = asset_name

        # asset_type 유추
        category = asset_data.get("category", "") or ""
        asset_data["asset_type"] = _derive_asset_type(category)

        # 상태 기본값
        asset_data["status"] = "active"
        asset_data["environment"] = "prod"

        # 파일 내 중복 검증
        if asset_name in seen_names:
            _append_error(
                errors,
                error_details,
                f"행 {excel_row}: 자산명 '{asset_name}'이(가) 행 {seen_names[asset_name]}과 중복됩니다.",
                sheet=INVENTORY_SHEET_NAME,
                row=excel_row,
                column="hostname",
                code="duplicate_asset_in_upload",
            )
            row_errors.append("중복 자산명")
        else:
            seen_names[asset_name] = excel_row

        # 필수 필드 검증: 최소한 category가 있어야 함
        if not category:
            warnings.append(f"행 {excel_row}: 대분류(category) 미입력")

        parsed_rows.append(asset_data)
        preview_rows.append(
            {
                "row_num": excel_row,
                "asset_name": asset_name,
                "asset_type": asset_data["asset_type"],
                "hostname": asset_data.get("hostname"),
                "vendor": asset_data.get("vendor"),
                "model": asset_data.get("model"),
                "serial_no": asset_data.get("serial_no"),
                "service_ip": asset_data.get("service_ip"),
                "mgmt_ip": asset_data.get("mgmt_ip"),
                "status": asset_data["status"],
                "errors": row_errors if row_errors else None,
            }
        )

    wb.close()

    if not parsed_rows:
        _append_error(
            errors, error_details, "데이터 행이 없습니다.", code="no_data_rows"
        )

    valid_count = len(parsed_rows) - sum(
        1 for pr in preview_rows if pr.get("errors")
    )

    return {
        "rows": parsed_rows,
        "preview_rows": preview_rows,
        "errors": errors,
        "error_details": error_details,
        "warnings": warnings,
        "total": len(parsed_rows),
        "valid_count": max(valid_count, 0),
    }


def import_inventory(
    db: Session,
    customer_id: int,
    parsed_rows: list[dict],
    current_user,
    on_duplicate: str = "skip",
) -> dict:
    """
    파싱된 자산 데이터를 DB에 저장.

    Args:
        on_duplicate: "skip" (기존 자산 건너뛰기) 또는 "overwrite" (기존 자산 덮어쓰기)

    Returns:
        {"created": int, "skipped": int, "errors": [], "error_details": []}
    """
    errors: list[str] = []
    error_details: list[dict] = []
    created = 0
    skipped = 0

    # 고객사 존재 확인
    if db.get(Customer, customer_id) is None:
        _append_error(errors, error_details, f"고객사(ID={customer_id})를 찾을 수 없습니다.", code="customer_not_found")
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}

    # 기존 자산명 조회 (고객사 범위)
    existing_assets = {
        a.asset_name: a
        for a in db.scalars(
            select(Asset).where(Asset.customer_id == customer_id)
        )
    }

    for row_data in parsed_rows:
        row_data["customer_id"] = customer_id
        asset_name = row_data.get("asset_name", "")
        if not asset_name:
            continue

        existing = existing_assets.get(asset_name)

        if existing:
            if on_duplicate == "skip":
                skipped += 1
                continue
            elif on_duplicate == "overwrite":
                # 기존 자산 업데이트
                for field, value in row_data.items():
                    if field in ("customer_id",) or value is None:
                        continue
                    setattr(existing, field, value)
                skipped += 1  # overwrite도 기존 건이므로 skipped에 포함
                continue

        # 새 자산 생성
        asset = Asset(**row_data)
        db.add(asset)
        created += 1

    if errors:
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}

    db.commit()

    return {"created": created, "skipped": skipped, "errors": [], "error_details": []}


# ──────────────────────────────────────────────────────────────
# IP Subnet (네트워크 대역) Import
# ──────────────────────────────────────────────────────────────

_SUBNET_HEADERS = ["name", "subnet", "netmask", "gateway", "vlan_id", "zone", "role", "region", "category", "description"]
_SUBNET_HEADER_LABELS = ["대역명", "Subnet (CIDR)", "Netmask", "Gateway", "VLAN", "Zone", "역할", "지역/층", "분류", "설명"]


def parse_subnet_sheet(file_bytes: bytes, customer_id: int) -> dict:
    """IP 대역 Excel 시트 파싱. Row 1=헤더, Row 2+= 데이터."""
    errors: list[str] = []
    error_details: list[dict] = []
    warnings: list[str] = []
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return {"rows": [], "preview_rows": [], "errors": errors, "error_details": error_details, "warnings": warnings, "total": 0, "valid_count": 0}

    ws = wb.active
    seen_subnets: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        name_val = _to_str(row[0].value) if len(row) > 0 else ""
        subnet_val = _to_str(row[1].value) if len(row) > 1 else ""
        if not name_val and not subnet_val:
            continue

        data: dict = {"customer_id": None}  # filled at import time
        row_errors: list[str] = []

        for i, field in enumerate(_SUBNET_HEADERS):
            val = row[i].value if len(row) > i else None
            data[field] = _to_str(val) or None

        if not data.get("name"):
            _append_error(errors, error_details, f"행 {excel_row}: 대역명 필수", sheet="IP대역", row=excel_row, column="name", code="missing_name")
            row_errors.append("대역명 누락")

        if not data.get("subnet"):
            _append_error(errors, error_details, f"행 {excel_row}: Subnet 필수", sheet="IP대역", row=excel_row, column="subnet", code="missing_subnet")
            row_errors.append("Subnet 누락")

        if data.get("role") is None:
            data["role"] = "service"

        # 중복 검증
        key = data.get("subnet", "")
        if key and key in seen_subnets:
            _append_error(errors, error_details, f"행 {excel_row}: Subnet '{key}' 중복", sheet="IP대역", row=excel_row, code="duplicate_subnet")
            row_errors.append("중복")
        elif key:
            seen_subnets[key] = excel_row

        parsed_rows.append(data)
        preview_rows.append({"row_num": excel_row, "name": data.get("name"), "subnet": data.get("subnet"), "vlan_id": data.get("vlan_id"), "zone": data.get("zone"), "role": data.get("role"), "errors": row_errors or None})

    wb.close()
    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")

    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))
    return {"rows": parsed_rows, "preview_rows": preview_rows, "errors": errors, "error_details": error_details, "warnings": warnings, "total": len(parsed_rows), "valid_count": max(valid_count, 0)}


def import_subnets(db: Session, customer_id: int, parsed_rows: list[dict], current_user, on_duplicate: str = "skip") -> dict:
    """파싱된 IP 대역 데이터를 DB에 저장."""
    errors: list[str] = []
    error_details: list[dict] = []
    created = 0
    skipped = 0

    if db.get(Customer, customer_id) is None:
        _append_error(errors, error_details, f"고객사(ID={customer_id})를 찾을 수 없습니다.", code="customer_not_found")
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}

    existing = {s.subnet: s for s in db.scalars(select(IpSubnet).where(IpSubnet.customer_id == customer_id))}

    for row_data in parsed_rows:
        row_data["customer_id"] = customer_id
        subnet_key = row_data.get("subnet", "")
        if not subnet_key:
            continue
        ex = existing.get(subnet_key)
        if ex:
            if on_duplicate == "skip":
                skipped += 1
            elif on_duplicate == "overwrite":
                for f, v in row_data.items():
                    if f == "customer_id" or v is None:
                        continue
                    setattr(ex, f, v)
                skipped += 1
            continue
        db.add(IpSubnet(**row_data))
        created += 1

    if errors:
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}
    db.commit()
    return {"created": created, "skipped": skipped, "errors": [], "error_details": []}


# ──────────────────────────────────────────────────────────────
# PortMap (포트맵) Import
# ──────────────────────────────────────────────────────────────

_PORTMAP_HEADERS = [
    "seq", "summary", "connection_type",
    "src_hostname", "src_ip", "src_slot", "src_port_name", "src_zone", "src_vlan",
    "dst_hostname", "dst_ip", "dst_slot", "dst_port_name", "dst_zone", "dst_vlan",
    "protocol", "port", "purpose", "status",
    "cable_type", "cable_speed", "cable_no", "note",
]
_PORTMAP_HEADER_LABELS = [
    "Seq", "요약", "접속유형",
    "출발 Hostname", "출발 IP", "출발 Slot", "출발 Port", "출발 Zone", "출발 VLAN",
    "도착 Hostname", "도착 IP", "도착 Slot", "도착 Port", "도착 Zone", "도착 VLAN",
    "Protocol", "Port No.", "용도", "상태",
    "케이블유형", "케이블속도", "케이블번호", "비고",
]

_PORTMAP_INT_FIELDS = {"seq", "port"}


def parse_portmap_sheet(file_bytes: bytes, customer_id: int) -> dict:
    """포트맵 Excel 시트 파싱. Row 1=헤더, Row 2+=데이터."""
    errors: list[str] = []
    error_details: list[dict] = []
    warnings: list[str] = []
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return {"rows": [], "preview_rows": [], "errors": errors, "error_details": error_details, "warnings": warnings, "total": 0, "valid_count": 0}

    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        # 빈 행 스킵
        src_host = _to_str(row[3].value) if len(row) > 3 else ""
        dst_host = _to_str(row[9].value) if len(row) > 9 else ""
        summary = _to_str(row[1].value) if len(row) > 1 else ""
        if not src_host and not dst_host and not summary:
            continue

        data: dict = {"customer_id": None}  # filled at import time
        row_errors: list[str] = []

        for i, field in enumerate(_PORTMAP_HEADERS):
            val = row[i].value if len(row) > i else None
            if field in _PORTMAP_INT_FIELDS:
                data[field] = _to_int(val)
            else:
                data[field] = _to_str(val) or None

        if not data.get("status"):
            data["status"] = "required"

        parsed_rows.append(data)
        preview_rows.append({
            "row_num": excel_row,
            "src_hostname": data.get("src_hostname"),
            "src_ip": data.get("src_ip"),
            "dst_hostname": data.get("dst_hostname"),
            "dst_ip": data.get("dst_ip"),
            "protocol": data.get("protocol"),
            "port": data.get("port"),
            "purpose": data.get("purpose"),
            "errors": row_errors or None,
        })

    wb.close()
    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")

    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))
    return {"rows": parsed_rows, "preview_rows": preview_rows, "errors": errors, "error_details": error_details, "warnings": warnings, "total": len(parsed_rows), "valid_count": max(valid_count, 0)}


def import_portmaps(db: Session, customer_id: int, parsed_rows: list[dict], current_user) -> dict:
    """파싱된 포트맵 데이터를 DB에 저장. 포트맵은 중복 개념이 약하므로 모두 신규 생성."""
    errors: list[str] = []
    error_details: list[dict] = []
    created = 0

    if db.get(Customer, customer_id) is None:
        _append_error(errors, error_details, f"고객사(ID={customer_id})를 찾을 수 없습니다.", code="customer_not_found")
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}

    for row_data in parsed_rows:
        row_data["customer_id"] = customer_id
        db.add(PortMap(**row_data))
        created += 1

    if errors:
        return {"created": 0, "skipped": 0, "errors": errors, "error_details": error_details}
    db.commit()
    return {"created": created, "skipped": 0, "errors": [], "error_details": []}


# ──────────────────────────────────────────────────────────────
# 샘플 양식 생성
# ──────────────────────────────────────────────────────────────

def build_sample_template(domain: str) -> bytes:
    """도메인별 Import 샘플 양식(빈 xlsx) 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if domain == "subnet":
        ws.title = "IP대역"
        ws.append(_SUBNET_HEADER_LABELS)
        # 샘플 행
        ws.append(["서버팜-Service", "10.0.1.0/24", "255.255.255.0", "10.0.1.1", "100", "Server-Zone", "service", "IDC-A 3F", "서버", "서버팜 Service 대역"])
    elif domain == "portmap":
        ws.title = "포트맵"
        ws.append(_PORTMAP_HEADER_LABELS)
        ws.append(["1", "WEB→WAS 연결", "내부", "WEB-01", "10.0.1.10", "", "eth0", "DMZ", "100", "WAS-01", "10.0.2.10", "", "eth0", "Service", "200", "TCP", "8080", "HTTP 요청", "required", "UTP", "1G", "", ""])
    elif domain == "inventory":
        ws.title = "01. Inventory"
        # 간소화된 헤더 (Row 1-2 그룹 헤더, Row 3 컬럼 헤더)
        ws.append([""] * 35)
        ws.append([""] * 35)
        headers = [
            "", "Seq", "센터", "운영구분", "장비관리번호", "Rack No.", "랙내 위치",
            "단계", "입고일", "대분류", "소분류", "제조사", "모델명", "Serial No.",
            "Hostname", "클러스터", "업무명", "Zone", "Service IP", "MGMT IP",
            "Size(unit)", "LC수량", "HA수량", "UTP수량", "전원수량", "전원Type",
            "Firmware Version", "자산 구분", "자산 번호", "도입년도", "관리부서",
            "담당자(정)", "담당자(부)", "유지보수업체", "비고",
        ]
        ws.append(headers)
    else:
        ws.title = "Sheet1"
        ws.append(["도메인이 올바르지 않습니다."])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
