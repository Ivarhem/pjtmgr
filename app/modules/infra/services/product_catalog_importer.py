"""Product Catalog Excel Import — hardware/SOFTWARE/MODEL catalog import helpers."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from uuid import uuid4

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.infra.models.product_catalog import ProductCatalog
from app.modules.infra.models.hardware_spec import HardwareSpec
from app.modules.infra.models.software_spec import SoftwareSpec
from app.modules.infra.models.model_spec import ModelSpec
from app.modules.infra.services.catalog_alias_service import resolve_vendor_canonical
from app.modules.infra.services.catalog_similarity_service import build_normalized_catalog_fields

# ── 공통 유틸 ──


def _to_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _to_int(val) -> int | None:
    s = _to_str(val)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _to_float(val) -> float | None:
    s = _to_str(val)
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_date(val):
    from datetime import date, datetime

    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = _to_str(val)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _append_error(
    errors: list[str],
    error_details: list[dict],
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    column: str | None = None,
    code: str | None = None,
) -> None:
    errors.append(message)
    error_details.append(
        {"message": message, "sheet": sheet, "row": row, "column": column, "code": code}
    )


def _make_empty_result() -> dict:
    return {
        "rows": [],
        "preview_rows": [],
        "errors": [],
        "error_details": [],
        "warnings": [],
        "total": 0,
        "valid_count": 0,
    }


# ── SPEC 시트 컬럼 정의 ──
# Row 1 = 헤더, Row 2+ = 데이터
_SPEC_HEADERS = [
    "vendor", "name", "product_type", "category",
    "size_unit", "width_mm", "height_mm", "depth_mm", "weight_kg",
    "power_count", "power_type", "power_watt",
    "cpu_summary", "memory_summary", "throughput_summary",
    "os_firmware", "spec_url", "reference_url",
]

_SPEC_HEADER_LABELS = [
    "제조사", "모델명", "제품유형", "분류",
    "Size(U)", "폭(mm)", "높이(mm)", "깊이(mm)", "무게(kg)",
    "전원수량", "전원유형", "전원(W)",
    "CPU 요약", "메모리 요약", "처리량 요약",
    "OS/FW", "스펙 URL", "참조 URL",
]

_SPEC_INT_FIELDS = {"size_unit", "width_mm", "height_mm", "depth_mm", "power_count", "power_watt"}
_SPEC_FLOAT_FIELDS = {"weight_kg"}

# ── EOSL 시트 컬럼 정의 ──
_EOSL_HEADERS = ["vendor", "name", "eos_date", "eosl_date", "eosl_note"]
_EOSL_HEADER_LABELS = ["제조사", "모델명", "EOS 일자", "EOSL 일자", "EOSL 비고"]
_EOSL_DATE_FIELDS = {"eos_date", "eosl_date"}

# ── SOFTWARE 시트 컬럼 정의 ──
_SOFTWARE_HEADERS = [
    "vendor", "name", "version", "product_type", "category",
    "edition", "license_type", "license_unit", "deployment_type",
    "runtime_env", "support_vendor", "reference_url", "architecture_note",
]
_SOFTWARE_HEADER_LABELS = [
    "제조사", "제품명", "버전", "상위분류", "분류",
    "에디션", "라이선스유형", "라이선스단위", "배포형태",
    "실행환경", "지원벤더", "참조URL", "비고",
]

# ── MODEL 시트 컬럼 정의 ──
_MODEL_HEADERS = [
    "vendor", "name", "version", "product_type", "category",
    "model_family", "modality", "deployment_scope", "context_window",
    "endpoint_format", "reference_url", "capability_note",
]
_MODEL_HEADER_LABELS = [
    "제공자", "모델명", "버전", "상위분류", "분류",
    "모델계열", "모달리티", "배포범위", "컨텍스트윈도우",
    "엔드포인트형식", "참조URL", "기능비고",
]
_MODEL_INT_FIELDS = {"context_window"}


# ── SPEC 파싱 ──


def parse_spec_sheet(file_bytes: bytes) -> dict:
    """SPEC 시트 파싱 → 프리뷰 데이터 반환."""
    result = _make_empty_result()
    errors = result["errors"]
    error_details = result["error_details"]
    warnings = result["warnings"]
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return result

    # SPEC 시트 찾기
    ws = None
    for name in wb.sheetnames:
        if "spec" in name.lower() or "스펙" in name:
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        warnings.append(f"'SPEC' 시트를 찾지 못해 첫 번째 시트('{wb.sheetnames[0]}')를 사용합니다.")

    if ws is None:
        _append_error(errors, error_details, "빈 파일입니다.", code="empty_file")
        wb.close()
        return result

    seen_keys: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        vendor_val = _to_str(row[0].value) if len(row) > 0 else ""
        name_val = _to_str(row[1].value) if len(row) > 1 else ""

        if not vendor_val and not name_val:
            continue

        data: dict = {}
        row_errors: list[str] = []

        for i, field in enumerate(_SPEC_HEADERS):
            val = row[i].value if len(row) > i else None
            if field in _SPEC_INT_FIELDS:
                data[field] = _to_int(val)
            elif field in _SPEC_FLOAT_FIELDS:
                data[field] = _to_float(val)
            else:
                data[field] = _to_str(val) or None

        # 필수 필드 검증
        if not data.get("vendor"):
            _append_error(errors, error_details, f"행 {excel_row}: 제조사 필수",
                          sheet="SPEC", row=excel_row, column="vendor", code="missing_vendor")
            row_errors.append("제조사 누락")

        if not data.get("name"):
            _append_error(errors, error_details, f"행 {excel_row}: 모델명 필수",
                          sheet="SPEC", row=excel_row, column="name", code="missing_name")
            row_errors.append("모델명 누락")

        if not data.get("category"):
            data["category"] = "기타"
            warnings.append(f"행 {excel_row}: 분류 미입력, '기타'로 지정")

        if not data.get("product_type"):
            data["product_type"] = "hardware"

        # 파일 내 중복 검증
        key = f"{data.get('vendor', '')}|{data.get('name', '')}"
        if key in seen_keys:
            _append_error(errors, error_details,
                          f"행 {excel_row}: '{data.get('vendor')} {data.get('name')}' 중복 (행 {seen_keys[key]})",
                          sheet="SPEC", row=excel_row, code="duplicate_product")
            row_errors.append("중복")
        else:
            seen_keys[key] = excel_row

        parsed_rows.append(data)
        preview_rows.append({
            "row_num": excel_row,
            "vendor": data.get("vendor"),
            "name": data.get("name"),
            "category": data.get("category"),
            "size_unit": data.get("size_unit"),
            "power_count": data.get("power_count"),
            "cpu_summary": data.get("cpu_summary"),
            "errors": row_errors or None,
        })

    wb.close()

    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")

    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))

    result["rows"] = parsed_rows
    result["preview_rows"] = preview_rows
    result["total"] = len(parsed_rows)
    result["valid_count"] = max(valid_count, 0)
    return result


def import_spec(
    db: Session,
    parsed_rows: list[dict],
    current_user,
    on_duplicate: str = "skip",
    import_batch_id: str | None = None,
) -> dict:
    """파싱된 SPEC 데이터를 DB에 upsert (product_catalog + hardware_specs)."""
    from app.modules.common.services import audit

    errors: list[str] = []
    error_details: list[dict] = []
    created = 0
    skipped = 0
    batch_id = import_batch_id or f"spec-{uuid4().hex[:12]}"
    imported_at = datetime.utcnow()

    # 기존 제품 조회
    existing = {
        (p.vendor, p.name): p
        for p in db.scalars(select(ProductCatalog))
    }

    # 분리할 필드 목록 (hardware_spec 전용)
    _SPEC_ONLY = {
        "size_unit", "width_mm", "height_mm", "depth_mm", "weight_kg",
        "power_count", "power_type", "power_watt",
        "cpu_summary", "memory_summary", "throughput_summary",
        "os_firmware", "spec_url",
    }

    for row_data in parsed_rows:
        vendor = resolve_vendor_canonical(db, row_data.get("vendor", "")) or ""
        row_data["vendor"] = vendor
        name = row_data.get("name", "")
        if not vendor or not name:
            continue

        # product_catalog 필드와 spec 필드 분리
        catalog_data = {k: v for k, v in row_data.items() if k not in _SPEC_ONLY}
        spec_data = {k: v for k, v in row_data.items() if k in _SPEC_ONLY}
        catalog_data.setdefault("source_name", "spec.xlsx")
        catalog_data.setdefault("source_confidence", "imported")
        catalog_data.setdefault("verification_status", "imported")
        catalog_data["import_batch_id"] = batch_id
        catalog_data["last_verified_at"] = imported_at
        catalog_data.pop("asset_type_key", None)
        catalog_data.update(build_normalized_catalog_fields(catalog_data.get("vendor"), catalog_data.get("name")))

        ex = existing.get((vendor, name))

        if ex:
            if on_duplicate == "skip":
                skipped += 1
                continue
            elif on_duplicate == "overwrite":
                # 제품 정보 업데이트
                for field, value in catalog_data.items():
                    if value is not None:
                        setattr(ex, field, value)
                # spec upsert
                _upsert_spec_data(db, ex.id, spec_data)
                skipped += 1
                continue

        # 새 제품 생성
        product = ProductCatalog(**catalog_data)
        db.add(product)
        db.flush()  # ID 확보

        # spec 생성
        has_spec = any(v is not None for v in spec_data.values())
        if has_spec:
            spec = HardwareSpec(product_id=product.id, **spec_data)
            db.add(spec)

        existing[(vendor, name)] = product
        created += 1

    if errors:
        return {
            "created": 0,
            "skipped": 0,
            "errors": errors,
            "error_details": error_details,
            "import_batch_id": batch_id,
        }

    audit.log(
        db, user_id=current_user.id, action="import", entity_type="product_catalog",
        entity_id=None, summary=f"제품 카탈로그 Import: {created}건 생성, {skipped}건 스킵",
        module="infra",
    )
    db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": [],
        "error_details": [],
        "import_batch_id": batch_id,
    }


def _upsert_spec_data(db: Session, product_id: int, spec_data: dict) -> None:
    """hardware_spec upsert (import 전용 헬퍼)."""
    spec = db.scalar(
        select(HardwareSpec).where(HardwareSpec.product_id == product_id)
    )
    if spec is None:
        has_spec = any(v is not None for v in spec_data.values())
        if has_spec:
            spec = HardwareSpec(product_id=product_id, **spec_data)
            db.add(spec)
    else:
        for field, value in spec_data.items():
            if value is not None:
                setattr(spec, field, value)


def _upsert_software_spec_data(db: Session, product_id: int, spec_data: dict) -> None:
    spec = db.scalar(
        select(SoftwareSpec).where(SoftwareSpec.product_id == product_id)
    )
    if spec is None:
        has_spec = any(v is not None for v in spec_data.values())
        if has_spec:
            spec = SoftwareSpec(product_id=product_id, **spec_data)
            db.add(spec)
    else:
        for field, value in spec_data.items():
            if value is not None:
                setattr(spec, field, value)


def _upsert_model_spec_data(db: Session, product_id: int, spec_data: dict) -> None:
    spec = db.scalar(
        select(ModelSpec).where(ModelSpec.product_id == product_id)
    )
    if spec is None:
        has_spec = any(v is not None for v in spec_data.values())
        if has_spec:
            spec = ModelSpec(product_id=product_id, **spec_data)
            db.add(spec)
    else:
        for field, value in spec_data.items():
            if value is not None:
                setattr(spec, field, value)


# ── EOSL 파싱 ──


def parse_eosl_sheet(file_bytes: bytes) -> dict:
    """EOSL 시트 파싱 → 프리뷰 데이터 반환."""
    result = _make_empty_result()
    errors = result["errors"]
    error_details = result["error_details"]
    warnings = result["warnings"]
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return result

    # EOSL 시트 찾기
    ws = None
    for name in wb.sheetnames:
        if "eosl" in name.lower() or "eol" in name.lower():
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        warnings.append(f"'EOSL' 시트를 찾지 못해 첫 번째 시트('{wb.sheetnames[0]}')를 사용합니다.")

    if ws is None:
        _append_error(errors, error_details, "빈 파일입니다.", code="empty_file")
        wb.close()
        return result

    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        vendor_val = _to_str(row[0].value) if len(row) > 0 else ""
        name_val = _to_str(row[1].value) if len(row) > 1 else ""

        if not vendor_val and not name_val:
            continue

        data: dict = {}
        row_errors: list[str] = []

        for i, field in enumerate(_EOSL_HEADERS):
            val = row[i].value if len(row) > i else None
            if field in _EOSL_DATE_FIELDS:
                data[field] = _to_date(val)
            else:
                data[field] = _to_str(val) or None

        if not data.get("vendor"):
            _append_error(errors, error_details, f"행 {excel_row}: 제조사 필수",
                          sheet="EOSL", row=excel_row, column="vendor", code="missing_vendor")
            row_errors.append("제조사 누락")

        if not data.get("name"):
            _append_error(errors, error_details, f"행 {excel_row}: 모델명 필수",
                          sheet="EOSL", row=excel_row, column="name", code="missing_name")
            row_errors.append("모델명 누락")

        parsed_rows.append(data)
        preview_rows.append({
            "row_num": excel_row,
            "vendor": data.get("vendor"),
            "name": data.get("name"),
            "eos_date": str(data["eos_date"]) if data.get("eos_date") else None,
            "eosl_date": str(data["eosl_date"]) if data.get("eosl_date") else None,
            "eosl_note": data.get("eosl_note"),
            "errors": row_errors or None,
        })

    wb.close()

    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")

    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))

    result["rows"] = parsed_rows
    result["preview_rows"] = preview_rows
    result["total"] = len(parsed_rows)
    result["valid_count"] = max(valid_count, 0)
    return result


def import_eosl(
    db: Session,
    parsed_rows: list[dict],
    current_user,
    on_duplicate: str = "skip",
    import_batch_id: str | None = None,
) -> dict:
    """파싱된 EOSL 데이터를 기존 제품에 업데이트. 제품이 없으면 스킵."""
    from app.modules.common.services import audit

    errors: list[str] = []
    error_details: list[dict] = []
    created = 0  # 여기서 created는 "업데이트된 건수"
    skipped = 0
    batch_id = import_batch_id or f"eosl-{uuid4().hex[:12]}"
    imported_at = datetime.utcnow()

    existing = {
        (p.vendor, p.name): p
        for p in db.scalars(select(ProductCatalog))
    }

    for row_data in parsed_rows:
        vendor = resolve_vendor_canonical(db, row_data.get("vendor", "")) or ""
        row_data["vendor"] = vendor
        name = row_data.get("name", "")
        if not vendor or not name:
            continue

        ex = existing.get((vendor, name))
        if ex is None:
            skipped += 1
            continue

        # EOS/EOSL 날짜 업데이트
        changed = False
        for field in ("eos_date", "eosl_date", "eosl_note"):
            val = row_data.get(field)
            if val is not None:
                setattr(ex, field, val)
                changed = True

        if changed:
            ex.source_name = ex.source_name or "spec.xlsx"
            ex.source_confidence = ex.source_confidence or "imported"
            ex.verification_status = "imported"
            ex.import_batch_id = batch_id
            ex.last_verified_at = imported_at

        if changed:
            created += 1
        else:
            skipped += 1

    if errors:
        return {
            "created": 0,
            "skipped": 0,
            "errors": errors,
            "error_details": error_details,
            "import_batch_id": batch_id,
        }

    audit.log(
        db, user_id=current_user.id, action="import", entity_type="product_catalog_eosl",
        entity_id=None, summary=f"EOSL Import: {created}건 업데이트, {skipped}건 스킵",
        module="infra",
    )
    db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": [],
        "error_details": [],
        "import_batch_id": batch_id,
    }


def parse_software_sheet(file_bytes: bytes) -> dict:
    """SOFTWARE 시트 파싱 → 프리뷰 데이터 반환."""
    result = _make_empty_result()
    errors = result["errors"]
    error_details = result["error_details"]
    warnings = result["warnings"]
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return result

    ws = None
    for name in wb.sheetnames:
        lowered = name.lower()
        if "software" in lowered or "소프트웨어" in name:
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        warnings.append(f"'SOFTWARE' 시트를 찾지 못해 첫 번째 시트('{wb.sheetnames[0]}')를 사용합니다.")

    if ws is None:
        _append_error(errors, error_details, "빈 파일입니다.", code="empty_file")
        wb.close()
        return result

    seen_keys: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        vendor_val = _to_str(row[0].value) if len(row) > 0 else ""
        name_val = _to_str(row[1].value) if len(row) > 1 else ""

        if not vendor_val and not name_val:
            continue

        data: dict = {}
        row_errors: list[str] = []

        for i, field in enumerate(_SOFTWARE_HEADERS):
            val = row[i].value if len(row) > i else None
            data[field] = _to_str(val) or None

        if not data.get("vendor"):
            _append_error(errors, error_details, f"행 {excel_row}: 제조사 필수",
                          sheet="SOFTWARE", row=excel_row, column="vendor", code="missing_vendor")
            row_errors.append("제조사 누락")
        if not data.get("name"):
            _append_error(errors, error_details, f"행 {excel_row}: 제품명 필수",
                          sheet="SOFTWARE", row=excel_row, column="name", code="missing_name")
            row_errors.append("제품명 누락")
        if not data.get("category"):
            data["category"] = "기타"
            warnings.append(f"행 {excel_row}: 분류 미입력, '기타'로 지정")
        if not data.get("product_type"):
            data["product_type"] = "software"

        key = f"{data.get('vendor', '')}|{data.get('name', '')}"
        if key in seen_keys:
            _append_error(errors, error_details,
                          f"행 {excel_row}: '{data.get('vendor')} {data.get('name')}' 중복 (행 {seen_keys[key]})",
                          sheet="SOFTWARE", row=excel_row, code="duplicate_product")
            row_errors.append("중복")
        else:
            seen_keys[key] = excel_row

        parsed_rows.append(data)
        preview_rows.append({
            "row_num": excel_row,
            "vendor": data.get("vendor"),
            "name": data.get("name"),
            "version": data.get("version"),
            "product_type": data.get("product_type"),
            "category": data.get("category"),
            "edition": data.get("edition"),
            "deployment_type": data.get("deployment_type"),
            "runtime_env": data.get("runtime_env"),
            "errors": row_errors or None,
        })

    wb.close()

    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")

    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))
    result["rows"] = parsed_rows
    result["preview_rows"] = preview_rows
    result["total"] = len(parsed_rows)
    result["valid_count"] = max(valid_count, 0)
    return result


def import_software(
    db: Session,
    parsed_rows: list[dict],
    current_user,
    on_duplicate: str = "skip",
    import_batch_id: str | None = None,
) -> dict:
    """파싱된 SOFTWARE 데이터를 DB에 upsert (product_catalog + software_spec)."""
    from app.modules.common.services import audit

    errors: list[str] = []
    error_details: list[dict] = []
    created = 0
    skipped = 0
    batch_id = import_batch_id or f"software-{uuid4().hex[:12]}"
    imported_at = datetime.utcnow()

    existing = {(p.vendor, p.name): p for p in db.scalars(select(ProductCatalog))}
    software_only = {
        "edition", "license_type", "license_unit", "deployment_type",
        "runtime_env", "support_vendor", "architecture_note",
    }

    for row_data in parsed_rows:
        vendor = resolve_vendor_canonical(db, row_data.get("vendor", "")) or ""
        row_data["vendor"] = vendor
        name = row_data.get("name", "")
        if not vendor or not name:
            continue

        catalog_data = {k: v for k, v in row_data.items() if k not in software_only}
        spec_data = {k: v for k, v in row_data.items() if k in software_only}
        catalog_data.setdefault("source_name", "software_import.xlsx")
        catalog_data.setdefault("source_confidence", "imported")
        catalog_data.setdefault("verification_status", "imported")
        catalog_data["import_batch_id"] = batch_id
        catalog_data["last_verified_at"] = imported_at
        catalog_data["product_type"] = catalog_data.get("product_type") or "software"
        catalog_data.pop("asset_type_key", None)
        catalog_data.update(build_normalized_catalog_fields(catalog_data.get("vendor"), catalog_data.get("name")))

        ex = existing.get((vendor, name))
        if ex:
            if on_duplicate == "skip":
                skipped += 1
                continue
            for field, value in catalog_data.items():
                if value is not None:
                    setattr(ex, field, value)
            _upsert_software_spec_data(db, ex.id, spec_data)
            skipped += 1
            continue

        product = ProductCatalog(**catalog_data)
        db.add(product)
        db.flush()
        if any(v is not None for v in spec_data.values()):
            db.add(SoftwareSpec(product_id=product.id, **spec_data))
        existing[(vendor, name)] = product
        created += 1

    audit.log(
        db, user_id=current_user.id, action="import", entity_type="product_catalog_software",
        entity_id=None, summary=f"소프트웨어 카탈로그 Import: {created}건 생성, {skipped}건 스킵",
        module="infra",
    )
    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors, "error_details": error_details, "import_batch_id": batch_id}


def parse_model_sheet(file_bytes: bytes) -> dict:
    """MODEL 시트 파싱 → 프리뷰 데이터 반환."""
    result = _make_empty_result()
    errors = result["errors"]
    error_details = result["error_details"]
    warnings = result["warnings"]
    parsed_rows: list[dict] = []
    preview_rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        _append_error(errors, error_details, f"파일을 읽을 수 없습니다: {e}", code="invalid_file")
        return result

    ws = None
    for name in wb.sheetnames:
        lowered = name.lower()
        if "model" in lowered or "llm" in lowered or "모델" in name:
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        warnings.append(f"'MODEL' 시트를 찾지 못해 첫 번째 시트('{wb.sheetnames[0]}')를 사용합니다.")

    if ws is None:
        _append_error(errors, error_details, "빈 파일입니다.", code="empty_file")
        wb.close()
        return result

    seen_keys: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        excel_row = row[0].row
        vendor_val = _to_str(row[0].value) if len(row) > 0 else ""
        name_val = _to_str(row[1].value) if len(row) > 1 else ""
        if not vendor_val and not name_val:
            continue

        data: dict = {}
        row_errors: list[str] = []
        for i, field in enumerate(_MODEL_HEADERS):
            val = row[i].value if len(row) > i else None
            if field in _MODEL_INT_FIELDS:
                data[field] = _to_int(val)
            else:
                data[field] = _to_str(val) or None

        if not data.get("vendor"):
            _append_error(errors, error_details, f"행 {excel_row}: 제공자 필수",
                          sheet="MODEL", row=excel_row, column="vendor", code="missing_vendor")
            row_errors.append("제공자 누락")
        if not data.get("name"):
            _append_error(errors, error_details, f"행 {excel_row}: 모델명 필수",
                          sheet="MODEL", row=excel_row, column="name", code="missing_name")
            row_errors.append("모델명 누락")
        if not data.get("category"):
            data["category"] = "기타"
            warnings.append(f"행 {excel_row}: 분류 미입력, '기타'로 지정")
        if not data.get("product_type"):
            data["product_type"] = "model"

        key = f"{data.get('vendor', '')}|{data.get('name', '')}"
        if key in seen_keys:
            _append_error(errors, error_details,
                          f"행 {excel_row}: '{data.get('vendor')} {data.get('name')}' 중복 (행 {seen_keys[key]})",
                          sheet="MODEL", row=excel_row, code="duplicate_product")
            row_errors.append("중복")
        else:
            seen_keys[key] = excel_row

        parsed_rows.append(data)
        preview_rows.append({
            "row_num": excel_row,
            "vendor": data.get("vendor"),
            "name": data.get("name"),
            "version": data.get("version"),
            "product_type": data.get("product_type"),
            "category": data.get("category"),
            "model_family": data.get("model_family"),
            "modality": data.get("modality"),
            "deployment_scope": data.get("deployment_scope"),
            "context_window": data.get("context_window"),
            "errors": row_errors or None,
        })

    wb.close()
    if not parsed_rows:
        _append_error(errors, error_details, "데이터 행이 없습니다.", code="no_data_rows")
    valid_count = len(parsed_rows) - sum(1 for p in preview_rows if p.get("errors"))
    result["rows"] = parsed_rows
    result["preview_rows"] = preview_rows
    result["total"] = len(parsed_rows)
    result["valid_count"] = max(valid_count, 0)
    return result


def import_model(
    db: Session,
    parsed_rows: list[dict],
    current_user,
    on_duplicate: str = "skip",
    import_batch_id: str | None = None,
) -> dict:
    """파싱된 MODEL 데이터를 DB에 upsert (product_catalog + model_spec)."""
    from app.modules.common.services import audit

    errors: list[str] = []
    error_details: list[dict] = []
    created = 0
    skipped = 0
    batch_id = import_batch_id or f"model-{uuid4().hex[:12]}"
    imported_at = datetime.utcnow()

    existing = {(p.vendor, p.name): p for p in db.scalars(select(ProductCatalog))}
    model_only = {
        "model_family", "modality", "deployment_scope",
        "context_window", "endpoint_format", "capability_note",
    }

    for row_data in parsed_rows:
        vendor = resolve_vendor_canonical(db, row_data.get("vendor", "")) or ""
        row_data["vendor"] = vendor
        name = row_data.get("name", "")
        if not vendor or not name:
            continue

        catalog_data = {k: v for k, v in row_data.items() if k not in model_only}
        spec_data = {k: v for k, v in row_data.items() if k in model_only}
        catalog_data.setdefault("source_name", "model_import.xlsx")
        catalog_data.setdefault("source_confidence", "imported")
        catalog_data.setdefault("verification_status", "imported")
        catalog_data["import_batch_id"] = batch_id
        catalog_data["last_verified_at"] = imported_at
        catalog_data["product_type"] = catalog_data.get("product_type") or "model"
        catalog_data.pop("asset_type_key", None)
        catalog_data.update(build_normalized_catalog_fields(catalog_data.get("vendor"), catalog_data.get("name")))

        ex = existing.get((vendor, name))
        if ex:
            if on_duplicate == "skip":
                skipped += 1
                continue
            for field, value in catalog_data.items():
                if value is not None:
                    setattr(ex, field, value)
            _upsert_model_spec_data(db, ex.id, spec_data)
            skipped += 1
            continue

        product = ProductCatalog(**catalog_data)
        db.add(product)
        db.flush()
        if any(v is not None for v in spec_data.values()):
            db.add(ModelSpec(product_id=product.id, **spec_data))
        existing[(vendor, name)] = product
        created += 1

    audit.log(
        db, user_id=current_user.id, action="import", entity_type="product_catalog_model",
        entity_id=None, summary=f"모델 카탈로그 Import: {created}건 생성, {skipped}건 스킵",
        module="infra",
    )
    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors, "error_details": error_details, "import_batch_id": batch_id}


# ── 샘플 양식 생성 ──


def build_spec_template() -> bytes:
    """SPEC Import 샘플 양식(빈 xlsx) 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPEC"
    ws.append(_SPEC_HEADER_LABELS)
    # 샘플 행
    ws.append([
        "Cisco", "Catalyst 9300-48P", "hardware", "네트워크",
        1, None, None, None, None,
        2, "AC", 715,
        None, None, "480 Gbps",
        "IOS-XE 17.x", None, None,
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_eosl_template() -> bytes:
    """EOSL Import 샘플 양식(빈 xlsx) 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EOSL"
    ws.append(_EOSL_HEADER_LABELS)
    ws.append(["Cisco", "Catalyst 9300-48P", "2028-12-31", "2031-12-31", "Last day of support"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_software_template() -> bytes:
    """SOFTWARE Import 샘플 양식(빈 xlsx) 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFTWARE"
    ws.append(_SOFTWARE_HEADER_LABELS)
    ws.append([
        "Oracle", "Oracle Database", "19c", "software", "middleware", "DBMS",
        "Enterprise", "subscription", "core", "on-prem",
        "Linux", "Oracle", "https://www.oracle.com/database/", None,
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_model_template() -> bytes:
    """MODEL Import 샘플 양식(빈 xlsx) 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MODEL"
    ws.append(_MODEL_HEADER_LABELS)
    ws.append([
        "OpenAI", "GPT-4.1", None, "model", "llm", "생성형AI",
        "GPT", "multimodal", "api", 1000000,
        "responses", "https://platform.openai.com/docs", None,
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
